# =================================================================================
# N.I.C.O.L.E. - Aplicaci√≥n de gesti√≥n de base de datos (archivo principal)
# ---------------------------------------------------------------------------------
# Descripci√≥n general (comentarios en espa√±ol):
# Este archivo contiene la aplicaci√≥n de escritorio usando customtkinter (CTk).
# Proporciona funciones para conectarse a MySQL, operaciones CRUD sobre tablas,
# utilidades de validaci√≥n y env√≠o de correos, y la interfaz gr√°fica completa.
#  - Peque√±as utilidades UI:
#      * mostrar_toast(texto, parent)
#      * _bind_responsive_wrap(win, labels, padding)
#      * _bind_responsive_button_text(win, button_text_pairs, padding, char_width)
#  - Widgets personalizados:
#      * PasswordEntry: entrada con bot√≥n para mostrar/ocultar contrase√±a
#  - Clase principal App(ctk.CTk): contiene toda la UI y la l√≥gica de la aplicaci√≥n
#      * login_usuario() - ventana de login
#      * _start_loading() / _stop_loading() - indicador animado de carga
#      * mostrar_bienvenida() - pantalla inicial despu√©s de login
#      * init_ui() - construye la interfaz principal (tabla, toolbar, botones, etc.)
#      * funciones de exportaci√≥n, registro de usuarios, recuperaci√≥n de contrase√±a,
#        edici√≥n directa de celdas, borrado de registros, historial, ayuda, etc.
#
# -------------------------
# IMPORTACI√ìN DE LIBRER√çAS
# -------------------------
# Se importan librer√≠as est√°ndar y de terceros necesarias para la app.

import hashlib
import re
import smtplib
import ssl
import random
from datetime import datetime, timedelta, timezone
from tkinter import messagebox, ttk, TclError, filedialog


import customtkinter as ctk
import pymysql
from PIL import Image, ImageTk
from pymysql.err import MySQLError, IntegrityError
import threading
import time

import csv
import openpyxl
from fpdf import FPDF
import os,sys
from dotenv import load_dotenv
import textwrap



# Carga las variables de entorno desde el archivo .env
load_dotenv()

# --------------------------------
# CONFIGURACI√ìN DE CREDENCIALES
# --------------------------------
# Carga las credenciales del servidor de correo electr√≥nico desde el archivo .env
EMAIL_HOST = os.getenv("EMAIL_HOST")
EMAIL_PORT = os.getenv("EMAIL_PORT")
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")

# -----------------------------------------------
# FUNCIONES DE VALIDACI√ìN Y GESTI√ìN DE LA BASE DE DATOS
# -----------------------------------------------

def resource_path(relative_path: str) -> str:
    """
    Devuelve la ruta absoluta de un recurso (archivo), considerando si el programa est√° empaquetado con PyInstaller.
    """
    try:
        # Cuando el programa est√° empaquetado con PyInstaller, usa la carpeta temporal
        base_path = sys._MEIPASS
    except AttributeError:
        # Si no existe, significa que estamos en desarrollo
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# Simple image cache to avoid reopening images repeatedly
_IMAGE_CACHE = {}
def load_ctk_image(name, size=None):
    key = (name, size)
    if key in _IMAGE_CACHE:
        return _IMAGE_CACHE[key]
    try:
        img = ctk.CTkImage(Image.open(resource_path(name)), size=size) if size else ctk.CTkImage(Image.open(resource_path(name)))
        _IMAGE_CACHE[key] = img
        return img
    except Exception:
        return None

def conectar_db():
    """
    Abre y retorna una conexi√≥n a la base de datos MySQL usando las variables de entorno. Devuelve None en caso de error.
    """
    try:
        return pymysql.connect(
            host=os.getenv("DB_HOST"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            database=os.getenv("DB_DATABASE"),
            port=int(os.getenv("DB_PORT")),
            ssl={"ca": resource_path("isrgrootx1.pem")}
        )
    except MySQLError as err:
        print(f"Error MySQL: No se pudo conectar a MySQL:\n{err}")
        return None

def init_db():
    """
    Crea las tablas necesarias para la aplicaci√≥n si no existen y asegura que exista un usuario administrador por defecto.
    """
    db = conectar_db()
    if not db:
        return

    try:
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historial (
                id INT AUTO_INCREMENT PRIMARY KEY,
                usuario VARCHAR(50),
                fecha DATETIME,
                accion VARCHAR(20),
                tabla VARCHAR(50),
                registro_id VARCHAR(50),
                valores_antes TEXT,
                valores_despues TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios_app (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50) UNIQUE,
                password_hash VARCHAR(255),
                email VARCHAR(255) UNIQUE,
                es_admin BOOLEAN
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS recuperacion_codigos (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50),
                email VARCHAR(255),
                codigo VARCHAR(6),
                expiracion DATETIME
            )
        """)

        # Crea el usuario admin si no existe
        cur.execute("SELECT * FROM usuarios_app WHERE username='admin'")
        if not cur.fetchone():
            h = hashlib.sha256("Forlyfe135@".encode()).hexdigest()
            cur.execute("INSERT INTO usuarios_app (username, password_hash, email, es_admin) VALUES (%s, %s, %s, %s)", ("admin", h, "nicole.informacion.1@gmail.com", True))

        db.commit()
    except MySQLError as err:
        messagebox.showerror("Error de Inicializaci√≥n", f"No se pudo inicializar la base de datos:\n{err}")
    finally:
        if db.open:
            cur.close()
            db.close()

def validar_email(email):
    """
    Valida sint√°cticamente una direcci√≥n de correo electr√≥nico.
    """
    regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.fullmatch(regex, email)

def validar_contrasena(password):
    """
    Verifica que una contrase√±a cumpla reglas b√°sicas de seguridad (longitud m√≠nima, may√∫scula y car√°cter especial).
    """
    if len(password) < 8:
        return "La contrase√±a debe tener al menos 8 caracteres."
    if not any(char.isupper() for char in password):
        return "La contrase√±a debe contener al menos una letra may√∫scula."
    if not re.search(r"[!@#$%^&*()_+={}\[\]|\\:;\"'<>,.?/`~]", password):
        return "La contrase√±a debe contener al menos un car√°cter especial."
    return None # La contrase√±a es v√°lida


def attach_password_checklist(parent, pass_widget):
    """
    Adjunta un checklist visual bajo el widget de contrase√±a.
    Los requisitos se muestran en rojo si no se cumplen y en verde si se cumplen.
    pass_widget puede ser un PasswordEntry (tiene .entry) o un CTkEntry.
    """
    try:
        # Determinar el widget de entrada real
        ent = getattr(pass_widget, 'entry', pass_widget)

        checklist_frame = ctk.CTkFrame(parent, fg_color="transparent")
        # Colocamos el frame justo despu√©s del widget de entrada usando pack/grid seg√∫n disponibilidad
        try:
            # intentamos pack si el widget usa pack
            checklist_frame.pack(fill="x", padx=16, pady=(6, 0))
        except Exception:
            try:
                checklist_frame.grid(row=99, column=0, sticky="ew", padx=6, pady=(6, 0))
            except Exception:
                checklist_frame.place(relx=0, rely=0)

        # Reglas y labels
        rules = [
            ("length", "Al menos 8 caracteres", lambda s: len(s) >= 8),
            ("upper", "Al menos una letra may√∫scula", lambda s: any(c.isupper() for c in s)),
            ("special", "Al menos un car√°cter especial", lambda s: re.search(r"[!@#$%^&*()_+={}\[\]|\\:;\"'<>,.?/`~]", s) is not None),
        ]

        labels = {}
        for key, text, _ in rules:
            lbl = ctk.CTkLabel(checklist_frame, text=text, font=ctk.CTkFont(size=11))
            lbl.pack(anchor="w")
            labels[key] = (lbl, _)

        def _update(ev=None):
            try:
                s = ent.get()
            except Exception:
                s = ""
            for key, (lbl, fn) in labels.items():
                try:
                    ok = fn(s)
                    lbl.configure(text_color="#4CAF50" if ok else "#E53935")
                except Exception:
                    lbl.configure(text_color="#E53935")

        # Vincular a eventos de escritura
        try:
            ent.bind('<KeyRelease>', _update)
        except Exception:
            pass

        # inicializar estado
        _update()
        return checklist_frame
    except Exception:
        return None

def enviar_email_cod(destinatario, codigo, tipo="recuperacion", username=None, email_destino=None):
    """
    Env√≠a un correo SMTP con un c√≥digo seg√∫n el tipo:
    - tipo='recuperacion' (por defecto)
    - tipo='registro' (confirmaci√≥n de cuenta)
    - tipo='cambio' (confirmaci√≥n de cambio de correo)

    username y email_destino son usados para personalizar la plantilla.
    """
    if not (EMAIL_HOST and EMAIL_PORT and EMAIL_USER and EMAIL_PASS and EMAIL_SENDER):
        messagebox.showwarning("Configuraci√≥n de Correo", "La configuraci√≥n de correo no est√° completa en el archivo .env.")
        return False
    
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(EMAIL_HOST, int(EMAIL_PORT), context=context) as server:
            server.login(EMAIL_USER, EMAIL_PASS.replace(" ", ""))

            # Seleccionar plantilla seg√∫n el tipo
            asunto = "C√≥digo de Recuperaci√≥n de Contrase√±a N.I.C.O.L.E"
            cuerpo = ""
            # Intentar obtener username si no fue pasado
            if not username:
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor()
                        cur.execute("SELECT username FROM usuarios_app WHERE email=%s", (destinatario,))
                        res = cur.fetchone()
                        if res:
                            username = res[0]
                except Exception:
                    username = None
                finally:
                    try:
                        if db and db.open:
                            cur.close(); db.close()
                    except Exception:
                        pass

            if tipo == 'registro':
                asunto = "Confirma tu cuenta en N.I.C.O.L.E"
                cuerpo = f"Hola {username if username else ''},\n\n¬°Bienvenido a N.I.C.O.L.E! Confirma tu cuenta ingresando este c√≥digo: {codigo}\n\nSi no solicitaste esto, ignora este mensaje.\n\nSaludos,\nEl equipo de N.I.C.O.L.E"
            elif tipo == 'cambio':
                asunto = "Confirma cambio de correo en N.I.C.O.L.E"
                cuerpo = f"Hola {username if username else ''},\n\nHemos recibido una solicitud para cambiar el correo de tu cuenta a {email_destino if email_destino else destinatario}. Para confirmar este cambio ingresa el c√≥digo: {codigo}\n\nSi no solicitaste este cambio, contacta con el administrador.\n\nSaludos,\nEl equipo de N.I.C.O.L.E"
            else:
                asunto = "C√≥digo de Recuperaci√≥n de Contrase√±a N.I.C.O.L.E"
                cuerpo = f"Hola {username if username else ''},\n\nHas solicitado un c√≥digo para recuperar tu contrase√±a.\nTu c√≥digo de verificaci√≥n es: {codigo}\n\nEste c√≥digo es v√°lido por 15 minutos. Si no solicitaste este cambio, ignora este mensaje.\n\nSaludos,\nEl equipo de N.I.C.O.L.E"

            msg = f"Subject: {asunto}\n\n{cuerpo}"
            server.sendmail(EMAIL_SENDER, destinatario, msg.encode('utf-8'))
        return True
    except Exception as e:
        messagebox.showerror("Error de Correo", f"No se pudo enviar el correo:\n{e}")
        return False

def registrar_historial(usuario, accion, tabla, registro_id, antes="", despues=""):
    """
    Inserta un registro de auditor√≠a en la tabla 'historial'. Enmascara campos sensibles cuando la tabla afectada es 'usuarios_app'.
    """
    db = conectar_db()
    if not db:
        return

    try:
        cur = db.cursor()
        # Enmascarar el ID del registro y la contrase√±a si la tabla es 'usuarios_app'
        if tabla == "usuarios_app":
            registro_id_hist = "***"
            
            # Enmascara el campo 'usuario' si est√° presente
            if "usuario" in despues:
                despues = re.sub(r"'usuario': '.*?'", "'usuario': '***'", despues)
            if "usuario" in antes:
                antes = re.sub(r"'usuario': '.*?'", "'usuario': '***'", antes)

            # Enmascara el campo 'correo' si est√° presente
            if "correo" in despues:
                despues = re.sub(r"'correo': '.*?'", "'correo': '***'", despues)
            if "correo" in antes:
                antes = re.sub(r"'correo': '.*?'", "'correo': '***'", antes)

            # Enmascara el campo 'password_hash' como ya lo ten√≠as
            if "password_hash" in despues:
                despues = re.sub(r"'password_hash': '.*?'", "'password_hash': '***'", despues)
            if "password_hash" in antes:
                antes = re.sub(r"'password_hash': '.*?'", "'password_hash': '***'", antes)
        else:
            registro_id_hist = str(registro_id)

        cur.execute("""
            INSERT INTO historial (usuario, fecha, accion, tabla, registro_id, valores_antes, valores_despues)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (usuario, datetime.now(), accion, tabla, registro_id_hist, antes, despues))
        db.commit()
    except MySQLError as err:
        print(f"Error al registrar en historial: {err}")
    finally:
        if db.open:
            cur.close()
            db.close()

def obtener_tablas():
    """
    Retorna la lista de tablas de la base de datos excluyendo las tablas internas usadas por la aplicaci√≥n.
    """
    db = conectar_db()
    if not db:
        return []

    cur = None
    try:
        # cache simple de tablas para reducir consultas repetidas
        if getattr(obtener_tablas, '_cache', None):
            return obtener_tablas._cache
        cur = db.cursor()
        cur.execute("SHOW TABLES")
        tablas = [r[0] for r in cur.fetchall() if r[0] not in ("historial", "usuarios_app", "recuperacion_codigos")]
        obtener_tablas._cache = tablas
        return tablas
    except MySQLError as err:
        messagebox.showerror("Error de Lectura", f"No se pudieron obtener las tablas:\n{err}")
        return []
    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception:
            pass
        try:
            if db.open:
                db.close()
        except Exception:
            pass

def _validar_nombre_tabla(tabla, lista_tablas_permitidas):
    """
    Comprueba que el nombre de tabla solicitado est√© en la lista de tablas permitidas (evita inyecci√≥n por nombres din√°micos).
    """
    if tabla in lista_tablas_permitidas:
        return True
    return False

def obtener_datos(tabla):
    """
    Recupera los nombres de columnas y todas las filas de la tabla indicada tras validar el nombre de la tabla.
    """
    db = conectar_db()
    if not db:
        return [], []

    tablas_validas = obtener_tablas()
    if not _validar_nombre_tabla(tabla, tablas_validas):
        messagebox.showerror("Error de Seguridad", "Nombre de tabla inv√°lido.")
        return [], []

    cur = None
    try:
        cur = db.cursor()
        # Por defecto traer s√≥lo las primeras 200 filas para acelerar la carga inicial
        cur.execute(f"SELECT * FROM `{tabla}` LIMIT 200")
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        return cols, rows
    except MySQLError as err:
        messagebox.showerror("Error de Lectura", f"No se pudieron obtener los datos de la tabla '{tabla}':\n{err}")
        return [], []
    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception:
            pass
        try:
            if db.open:
                db.close()
        except Exception:
            pass

def mostrar_toast(texto, parent):
    """
    Muestra un peque√±o Toplevel temporal con un mensaje (notificaci√≥n tipo toast).
    """
    toast = ctk.CTkToplevel(parent)
    toast.overrideredirect(True)
    
    x = parent.winfo_rootx() + parent.winfo_width() - 310
    y = parent.winfo_rooty() + parent.winfo_height() - 110
    toast.geometry(f"300x50+{x}+{y}")
    
    ctk.CTkLabel(toast, text=texto, font=ctk.CTkFont(size=14)).pack(expand=True, fill="both", padx=10, pady=5)
    toast.attributes("-topmost", True)
    toast.after(2500, toast.destroy)


def _bind_responsive_wrap(win, labels, padding=40):
    """
    Vincula el evento de redimensionado para ajustar autom√°ticamente el wraplength de una lista de etiquetas (para interfaces responsivas).
    """
    def _on_resize(event=None):
        try:
            w = max(150, win.winfo_width() - padding)
        except Exception:
            w = 300
        for lab in labels:
            try:
                lab.configure(wraplength=w)
            except Exception:
                pass

    win.bind('<Configure>', _on_resize)

def _bind_responsive_button_text(win, button_text_pairs, padding=20, char_width=7):
    """
    Ajusta din√°micamente el texto de los botones para que se envuelva en varias l√≠neas cuando el ancho sea peque√±o. Detecta iconos iniciales (emoji) y los coloca sobre el texto si se requiere.

    button_text_pairs: iterable de (widget_boton, texto_original)
    """
    def _on_resize(event=None):
    # Para cada bot√≥n: calcular el ancho disponible en su padre y envolver el texto
        for btn, orig in button_text_pairs:
            try:
                # preferir el ancho propio del bot√≥n si est√° disponible (m√°s preciso)
                try:
                    avail = btn.winfo_width() - padding
                except Exception:
                    avail = 0

                if not avail or avail <= 10:
                    parent = getattr(btn, 'master', win)
                    try:
                        avail = parent.winfo_width() - padding
                    except Exception:
                        avail = 0

                if not avail or avail <= 0:
                    avail = max(50, win.winfo_width() - padding)

                # estimar el n√∫mero de caracteres que caben en una l√≠nea
                # (ajusta char_width para controlar el punto de quiebre)
                chars = max(1, int(avail / max(4, char_width)))

                # Si el texto original comienza con un icono (emoji/s√≠mbolo) seguido
                # de un espacio, separarlo en `icon` y el resto de texto usando el
                # primer espacio. Maneja emoji compuestos (por ejemplo üóëÔ∏è) que pueden
                # contener selectores y no tener un espacio fijo en un √≠ndice exacto.
                icon = None
                rest = orig
                try:
                    parts = orig.split(' ', 1)
                    if len(parts) > 1:
                        icon = parts[0]
                        rest = parts[1]
                except Exception:
                    icon = None
                    rest = orig

                # Si hay suficiente espacio, mantener el texto en una sola l√≠nea
                if chars >= len(orig.replace('\n', '')):
                    out = orig
                else:
                    # envolver la parte textual y, si existe un icono, colocarlo encima
                    wrapped_lines = textwrap.wrap(rest, width=chars) if rest else []
                    if icon:
                        if wrapped_lines:
                            out = icon + "\n" + "\n".join(wrapped_lines)
                        else:
                            out = icon
                    else:
                        out = "\n".join(textwrap.wrap(orig, width=chars))

                # aplicar el texto envuelto y solicitar centrado
                btn.configure(text=out)
                try:
                    btn.configure(justify="center")
                except Exception:
                    pass
            except Exception:
                # ignore layout issues during startup
                pass

    # Vincular al evento de redimensionado de la ventana y al padre de cada
    # bot√≥n para actualizaciones m√°s granulares
    win.bind('<Configure>', _on_resize)
    for btn, _ in button_text_pairs:
        try:
            btn.master.bind('<Configure>', _on_resize)
        except Exception:
            pass
    # run once to initialize button labels
    try:
        _on_resize()
    except Exception:
        pass

# -----------------------------------------------
# CLASES DE LA INTERFAZ DE USUARIO
# -----------------------------------------------

class PasswordEntry(ctk.CTkFrame):
    """
    Widget de entrada de contrase√±a con un bot√≥n para mostrar/ocultar la contrase√±a.
    Es una mejora sobre el widget de entrada est√°ndar de customtkinter.
    """
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)

        self.entry = ctk.CTkEntry(self, show="*")
        self.entry.grid(row=0, column=0, sticky="ew")

        try:
            self.eye_open_image = load_ctk_image("eye_open.png", size=(20, 20))
            self.eye_closed_image = load_ctk_image("eye_closed.png", size=(20, 20))
        except Exception:
            self.eye_open_image = None
            self.eye_closed_image = None
            messagebox.showwarning("Icono no encontrado", "Aseg√∫rate de tener 'eye_open.png' y 'eye_closed.png' en la carpeta.")

        self.show_pass_button = ctk.CTkButton(
            self,
            image=self.eye_open_image,
            text="",
            width=30,
            command=self.toggle_show_password,
            fg_color="transparent",
        )
        self.show_pass_button.grid(row=0, column=1, padx=(5, 0))

        self.is_password_visible = False

    def toggle_show_password(self):
        """Alterna la visibilidad de la contrase√±a."""
        if self.is_password_visible:
            self.entry.configure(show="*")
            self.show_pass_button.configure(image=self.eye_open_image)
        else:
            self.entry.configure(show="")
            self.show_pass_button.configure(image=self.eye_closed_image)
        self.is_password_visible = not self.is_password_visible

    def get(self):
        """Retorna el texto del campo de entrada."""
        return self.entry.get()
    
    def insert(self, index, string):
        """Inserta texto en el campo de entrada."""
        self.entry.insert(index, string)

    def pack(self, **kwargs):
        # Let geometry managers control sizing; entry expands via grid
        return super().pack(**kwargs)

    def bind(self, sequence, func, add='+'):
        self.entry.bind(sequence, func, add=add)

class App(ctk.CTk):
    """
    Clase principal que encapsula toda la aplicaci√≥n, manejando la UI y eventos.
    Contiene la l√≥gica para el login, la visualizaci√≥n de tablas y la interacci√≥n con la base de datos.
    """
    def __init__(self):
        super().__init__()
        self.withdraw()
        init_db()
        self.usuario = None
        self.is_admin = False
        self.login_usuario()

    def login_usuario(self):
        """Muestra la ventana de login al iniciar la aplicaci√≥n."""
        win = ctk.CTkToplevel(self)
        win.title("üîê Login")
        # allow the window to size dynamically; set a sensible minimum
        win.minsize(480, 320)
        win.resizable(True, True)
        win.protocol("WM_DELETE_WINDOW", self.destroy)
        win.transient(self)
        win.after(100, win.grab_set)

        win.grid_rowconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)
        win.grid_rowconfigure(2, weight=1)
        win.grid_columnconfigure(0, weight=1)

        frame = ctk.CTkFrame(win)
        frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_rowconfigure(2, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(frame, text="Usuario:").grid(row=0, column=0, sticky="ew", pady=(20, 5))
        entry_user = ctk.CTkEntry(frame, placeholder_text="nombre de usuario")
        entry_user.grid(row=1, column=0, sticky="ew", padx=10)
        ctk.CTkLabel(frame, text="Contrase√±a:").grid(row=2, column=0, sticky="ew", pady=5)
        entry_pass_frame = PasswordEntry(frame)
        entry_pass_frame.grid(row=3, column=0, sticky="ew", padx=10)

        def intentar(event=None):
            u = entry_user.get()
            p = entry_pass_frame.get()
            if not u or not p:
                mostrar_toast("Usuario y contrase√±a requeridos", win)
                return

            h = hashlib.sha256(p.encode()).hexdigest()
            db = conectar_db()
            if not db: return

            try:
                cur = db.cursor()
                cur.execute("SELECT es_admin FROM usuarios_app WHERE username=%s AND password_hash=%s", (u, h))
                res = cur.fetchone()
                if res:
                    self.usuario = u
                    self.is_admin = bool(res[0])
                    win.destroy()
                    self.mostrar_bienvenida()
                else:
                    mostrar_toast("Credenciales inv√°lidas", win)
            except MySQLError as err:
                messagebox.showerror("Error de Login", f"Ocurri√≥ un error:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=4, column=0, sticky="ew", pady=12, padx=10)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(btn_frame, text="Entrar", command=intentar).grid(row=0, column=0, sticky="ew", padx=5)
        ctk.CTkButton(btn_frame, text="Olvid√© mi contrase√±a", command=self.recuperar_contrasena, fg_color="gray", hover_color="darkgray").grid(row=0, column=1, sticky="ew", padx=5)

        entry_user.bind('<Return>', intentar)
        entry_pass_frame.bind('<Return>', intentar)

    # --- Loading indicator helpers ---
    def _start_loading(self, parent=None, text="Cargando..."):
        """Show a small animated loading label (spinner) overlaying the parent or main window.
        Prefer an in-window overlay frame (so it's always visible), fallback to Toplevel on error.
        """
        if not parent:
            parent = self
        # if already showing, skip
        try:
            if hasattr(self, '_loading_widget') and self._loading_widget.winfo_exists():
                return
        except Exception:
            pass

        # Try in-window overlay first
        try:
            parent.update_idletasks()
            w = 260
            h = 64
            # create overlay frame centered in parent
            self._loading_widget = ctk.CTkFrame(parent, corner_radius=8, fg_color="#2b2f31")
            self._loading_widget.place(relx=0.5, rely=0.5, anchor='center', width=w, height=h)
            self._loading_widget.lift()

            self._loading_label = ctk.CTkLabel(self._loading_widget, text=text, font=ctk.CTkFont(size=12))
            self._loading_label.pack(side='left', padx=(12,6), pady=8)
            self._loading_spinner = ctk.CTkLabel(self._loading_widget, text="", font=ctk.CTkFont(size=14))
            self._loading_spinner.pack(side='right', padx=(6,12))

            self._loading_frames = ['‚£æ','‚£Ω','‚£ª','‚¢ø','‚°ø','‚£ü','‚£Ø','‚£∑']
            self._loading_frame_index = 0

            def _animate():
                try:
                    self._loading_spinner.configure(text=self._loading_frames[self._loading_frame_index])
                    self._loading_frame_index = (self._loading_frame_index + 1) % len(self._loading_frames)
                    self._loading_widget.after(120, _animate)
                except Exception:
                    return

            _animate()
            return
        except Exception:
            # fallback to Toplevel if in-window overlay fails
            pass

        try:
            self._loading_widget = ctk.CTkToplevel(parent)
            self._loading_widget.overrideredirect(True)
            self._loading_widget.attributes('-topmost', True)
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            pw = parent.winfo_width()
            ph = parent.winfo_height()
            w = 220
            h = 50
            try:
                self._loading_widget.geometry(f"{w}x{h}+{px + max(0,(pw - w)//2)}+{py + max(0,(ph - h)//2)}")
            except Exception:
                self._loading_widget.geometry(f"{w}x{h}")

            frame = ctk.CTkFrame(self._loading_widget)
            frame.pack(expand=True, fill='both')
            self._loading_label = ctk.CTkLabel(frame, text=text, font=ctk.CTkFont(size=12))
            self._loading_label.pack(side='left', padx=(12,6), pady=8)
            self._loading_spinner = ctk.CTkLabel(frame, text="", font=ctk.CTkFont(size=14))
            self._loading_spinner.pack(side='right', padx=(6,12))

            self._loading_frames = ['‚£æ','‚£Ω','‚£ª','‚¢ø','‚°ø','‚£ü','‚£Ø','‚£∑']
            self._loading_frame_index = 0

            def _animate2():
                try:
                    self._loading_spinner.configure(text=self._loading_frames[self._loading_frame_index])
                    self._loading_frame_index = (self._loading_frame_index + 1) % len(self._loading_frames)
                    self._loading_widget.after(120, _animate2)
                except Exception:
                    return

            _animate2()
        except Exception:
            pass

    def _stop_loading(self):
        try:
            if hasattr(self, '_loading_widget') and self._loading_widget.winfo_exists():
                # if placed in window, use place_forget; else destroy
                try:
                    self._loading_widget.place_forget()
                except Exception:
                    try:
                        self._loading_widget.destroy()
                    except Exception:
                        pass
        except Exception:
            pass

    def mostrar_bienvenida(self):
        """Muestra una ventana de bienvenida despu√©s de un login exitoso."""
        self.withdraw()
        win = ctk.CTkToplevel(self)
        win.title("üéâ Bienvenida")
        win.minsize(360, 220)
        win.resizable(True, True)
        win.protocol("WM_DELETE_WINDOW", self.destroy)
        win.transient(self)
        win.after(10, win.grab_set)

        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(0, weight=1)
        frame = ctk.CTkFrame(win)
        frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_rowconfigure(2, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(frame, text=f"¬°Bienvenido, {self.usuario}!", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, sticky="ew", pady=(30, 10))
        ctk.CTkLabel(frame, text="Has iniciado sesi√≥n correctamente.", justify="center").grid(row=1, column=0, sticky="ew")

        def continuar():
            win.destroy()
            self.deiconify()
            self.init_ui()

        ctk.CTkButton(frame, text="üëâ Continuar", command=continuar).grid(row=2, column=0, sticky="ew", pady=30)
        win.bind('<Return>', lambda event: continuar())

    def cerrar_sesion(self):
        """Cierra la sesi√≥n del usuario y vuelve a la ventana de login."""
        self.withdraw()
        for widget in self.winfo_children():
            widget.destroy()
        self.usuario = None
        self.is_admin = False
        self.login_usuario()

    def init_ui(self):
        """Inicializa la interfaz principal de la aplicaci√≥n."""
        self.title(f"N.I.C.O.L.E ‚Äì Sesi√≥n de {self.usuario}")
        # Start maximized where available but allow restore
        try:
            self.state('zoomed')
        except Exception:
            pass

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=10)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        try:
            self.brain_icon = load_ctk_image("brain_icon.png", size=(40, 40))
            self.help_icon = load_ctk_image("help_icon.png", size=(25, 25))
        except FileNotFoundError:
            self.brain_icon = None
            self.help_icon = None
            messagebox.showwarning("Iconos no encontrados", "Aseg√∫rate de tener los iconos necesarios.")

        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=8)
        top_frame.grid_columnconfigure(0, weight=1)
        top_frame.grid_columnconfigure(1, weight=10)
        top_frame.grid_columnconfigure(2, weight=2)
        top_frame.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(top_frame, text="N.I.C.O.L.E", image=self.brain_icon, compound="left", font=ctk.CTkFont(size=22, weight="bold")).grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        # search entry should expand naturally
        self.busc = ctk.CTkEntry(top_frame, placeholder_text="üîç Buscar en la tabla actual‚Ä¶")
        self.busc.grid(row=0, column=1, sticky="ew", padx=6)
        self.busc.bind("<KeyRelease>", self.buscar)

        frame_selector = ctk.CTkFrame(top_frame, fg_color="transparent")
        frame_selector.grid(row=0, column=2, padx=6, sticky="ew")
        frame_selector.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(frame_selector, text="Tabla:").grid(row=0, column=0, sticky="ew", padx=(0, 5))
        tablas = obtener_tablas()
        self.combo = ctk.CTkOptionMenu(frame_selector, values=tablas if tablas else ["No hay tablas"], command=self.cargar_tabla)
        self.combo.grid(row=0, column=1, sticky="ew")

        # Bot√≥n de Ayuda en la esquina superior derecha
        self.btn_help = ctk.CTkButton(top_frame, text="", image=self.help_icon, width=40, command=self.mostrar_ayuda)
        self.btn_help.grid(row=0, column=3, padx=6, pady=6, sticky="e")

        ttk_frame = ctk.CTkFrame(self)
        ttk_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=(0, 8))
        ttk_frame.grid_rowconfigure(0, weight=1)
        ttk_frame.grid_columnconfigure(0, weight=1)

        # Estilos de la tabla (ttk.Treeview)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2a2d2e", foreground="white", rowheight=25, fieldbackground="#343638", borderwidth=0)
        style.map('Treeview', background=[('selected', '#245d81')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat", font=('Calibri', 10, 'bold'))
        style.map("Treeview.Heading", background=[('active', '#343638')])

        self.tree = ttk.Treeview(ttk_frame, style="Treeview")
        vsb = ttk.Scrollbar(ttk_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(ttk_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<Double-1>", self.editar_celda)

        bottom_frame = ctk.CTkFrame(self)
        bottom_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=8)
        bottom_frame.grid_columnconfigure(0, weight=2)
        bottom_frame.grid_columnconfigure(2, weight=2)
        bottom_frame.grid_columnconfigure(4, weight=2)
        # Botones de acci√≥n (A√±adir, Borrar)
        frame_acciones = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_acciones.grid(row=0, column=0, sticky="ew")
        frame_acciones.grid_columnconfigure(0, weight=1)
        frame_acciones.grid_columnconfigure(1, weight=1)
        self.btn_add = ctk.CTkButton(frame_acciones, text="‚ûï A√±adir", command=self.abrir_form)
        self.btn_add.grid(row=0, column=0, sticky="ew", padx=5)
        # Make the delete button center its content and allow vertical wrapping
        self.btn_del = ctk.CTkButton(
            frame_acciones,
            text="üóëÔ∏è Borrar",
            fg_color="#D32F2F",
            hover_color="#B71C1C",
            command=self.eliminar_registro,
            anchor='center'
        )
        # add a bit more internal padding so the icon doesn't feel cramped when
        # it wraps above the label on narrow windows
        try:
            self.btn_del.configure(padx=8, pady=6)
        except Exception:
            pass
        self.btn_del.grid(row=0, column=1, sticky="ew", padx=5)

        # Botones de gesti√≥n (Recargar, Historial, Exportar)
        frame_gestion = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_gestion.grid(row=0, column=2, sticky="ew")
        frame_gestion.grid_columnconfigure(0, weight=1)
        frame_gestion.grid_columnconfigure(1, weight=1)
        frame_gestion.grid_columnconfigure(2, weight=1)
        # keep references so we can update their text responsively
        recargar_btn = ctk.CTkButton(frame_gestion, text="üîÑ Recargar", command=self.cargar_tabla_actual)
        recargar_btn.grid(row=0, column=0, sticky="ew", padx=5)
        historial_btn = ctk.CTkButton(frame_gestion, text="üïµÔ∏è Historial", command=self.ver_historial)
        historial_btn.grid(row=0, column=1, sticky="ew", padx=5)
        export_btn = ctk.CTkButton(frame_gestion, text="üì¶ Exportar", command=self.iniciar_exportacion)
        export_btn.grid(row=0, column=2, sticky="ew", padx=5)

        # Botones de sesi√≥n (Usuarios, Cambiar datos, Salir)
        frame_sesion = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_sesion.grid(row=0, column=4 , sticky="ew")
        frame_sesion.grid_columnconfigure(0, weight=1)
        frame_sesion.grid_columnconfigure(1, weight=1)
        frame_sesion.grid_columnconfigure(2, weight=1)
        self.btn_user = ctk.CTkButton(frame_sesion, text="üë§ Nuevos usuarios", command=self.registrar_usuario)
        self.btn_user.grid(row=0, column=0, sticky="ew", padx=5)
        self.btn_change_data = ctk.CTkButton(frame_sesion, text="‚öôÔ∏è Cambiar mis datos", command=self.cambiar_datos)
        self.btn_change_data.grid(row=0, column=1, sticky="ew", padx=5)
        exit_btn = ctk.CTkButton(frame_sesion, text="üö™ Salir", command=self.cerrar_sesion)
        exit_btn.grid(row=0, column=2, sticky="ew", padx=5)

        # Deshabilita los botones de administrador si el usuario no es admin
        if not self.is_admin:
            self.btn_del.configure(state="disabled")
            self.btn_user.configure(state="disabled")
            self.btn_add.configure(state="disabled")
            self.tree.unbind("<Double-1>")

        if tablas:
            self.combo.set(tablas[0])
            self.cargar_tabla(tablas[0])

        # Make bottom toolbar buttons wrap their labels when space is constrained
        try:
            _bind_responsive_button_text(self, [
                (self.btn_add, '‚ûï A√±adir'),
                (self.btn_del, 'üóëÔ∏è Borrar'),
                (recargar_btn, 'üîÑ Recargar'),
                (historial_btn, 'üïµÔ∏è Historial'),
                (export_btn, 'üì¶ Exportar'),
                (self.btn_user, 'üë§ Nuevos usuarios'),
                (self.btn_change_data, '‚öôÔ∏è Cambiar mis datos'),
                (exit_btn, 'üö™ Salir'),
            ])
        except Exception:
            pass

    def iniciar_exportacion(self):
        """Inicia el proceso de exportaci√≥n de datos a diferentes formatos."""
        if not hasattr(self, "table") or self.table == "No hay tablas":
            return mostrar_toast("Primero selecciona una tabla con datos.", self)
        
        file_types = [
            ("Archivos Excel", "*.xlsx"),
            ("Archivos PDF", "*.pdf"),
            ("Archivos CSV", "*.csv")
        ]
        
        filepath = filedialog.asksaveasfilename(
            filetypes=file_types,
            defaultextension=".xlsx",
            initialfile=f"export_{self.table}_{datetime.now():%Y-%m-%d}",
            title="Exportar datos de la tabla"
        )

        if not filepath:
            return

        if filepath.endswith('.xlsx'):
            self._exportar_excel(filepath)
        elif filepath.endswith('.csv'):
            self._exportar_csv(filepath)
        elif filepath.endswith('.pdf'):
            self._exportar_pdf(filepath)

    def _exportar_csv(self, filepath):
        """Funci√≥n auxiliar para exportar datos a un archivo CSV."""
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(self.col)
                for item_id in self.tree.get_children():
                    writer.writerow(self.tree.item(item_id)['values'])
            messagebox.showinfo("√âxito", "CSV exportado con √©xito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportaci√≥n", f"No se pudo exportar a CSV:\n{e}")

    def _exportar_excel(self, filepath):
        """Funci√≥n auxiliar para exportar datos a un archivo Excel (.xlsx)."""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.table
            sheet.append(self.col)
            for item_id in self.tree.get_children():
                sheet.append(self.tree.item(item_id)['values'])
            workbook.save(filepath)
            messagebox.showinfo("√âxito", "Excel exportado con √©xito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportaci√≥n", f"No se pudo exportar a Excel:\n{e}")

    def _exportar_pdf(self, filepath):
        """Funci√≥n auxiliar para exportar datos a un archivo PDF."""
        try:
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, f'Reporte de Tabla: {self.table}', 0, 1, 'C')
            pdf.ln(5)

            pdf.set_font("Arial", 'B', 8)
            num_cols = len(self.col)
            page_width = pdf.w - 2 * pdf.l_margin
            col_width = page_width / num_cols if num_cols > 0 else 10

            pdf.set_fill_color(224, 224, 224)
            for header in self.col:
                pdf.cell(col_width, 8, str(header), 1, 0, 'C', fill=True)
            pdf.ln()

            pdf.set_font("Arial", '', 7)
            pdf.set_fill_color(255, 255, 255)
            fill = False
            for item_id in self.tree.get_children():
                row = self.tree.item(item_id)['values']
                for item in row:
                    pdf.cell(col_width, 8, str(item), 1, 0, 'L', fill=fill)
                pdf.ln()
                fill = not fill

            pdf.output(filepath)
            messagebox.showinfo("√âxito", "PDF exportado con √©xito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportaci√≥n", f"No se pudo exportar a PDF:\n{e}")
            
    def registrar_usuario(self):
        """Muestra una ventana para que un administrador pueda crear nuevos usuarios."""
        if not self.is_admin: return
        win = ctk.CTkToplevel(self)
        win.title("‚ûï Nuevo Usuario")
        win.minsize(380, 320)
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        lbl_user = ctk.CTkLabel(win, text="Nombre de Usuario:")
        lbl_user.pack(pady=(14, 4), anchor="w", padx=16)
        e_user = ctk.CTkEntry(win)
        e_user.pack(padx=16, fill="x")

        lbl_email = ctk.CTkLabel(win, text="Correo Electr√≥nico:")
        lbl_email.pack(pady=(8, 4), anchor="w", padx=16)
        e_email = ctk.CTkEntry(win)
        e_email.pack(padx=16, fill="x")

        lbl_pass = ctk.CTkLabel(win, text="Contrase√±a:")
        lbl_pass.pack(pady=(8, 4), anchor="w", padx=16)
        e_pass_frame = PasswordEntry(win)
        e_pass_frame.pack(padx=16, fill="x")
        # Adjuntar checklist de contrase√±a
        try:
            attach_password_checklist(win, e_pass_frame)
        except Exception:
            pass

        val_admin = ctk.CTkCheckBox(win, text="Otorgar privilegios de Administrador")
        val_admin.pack(pady=10)

        def guardar_user():
            """Inicia el flujo de creaci√≥n: env√≠a un c√≥digo al correo y solicita confirmaci√≥n."""
            u = e_user.get().strip()
            p = e_pass_frame.get().strip()
            e = e_email.get().strip()

            if not u or not p or not e:
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)

            if not validar_email(e):
                return messagebox.showwarning("Correo inv√°lido", "Formato de correo inv√°lido.", parent=win)

            validation_error = validar_contrasena(p)
            if validation_error:
                return messagebox.showwarning("Contrase√±a d√©bil", validation_error, parent=win)

            # Generar c√≥digo temporal y guardarlo
            codigo = "".join(random.choices("0123456789", k=6))
            expiracion = datetime.now(timezone.utc) + timedelta(minutes=15)

            db = conectar_db()
            if not db:
                return
            try:
                cur = db.cursor()
                # Guardar c√≥digo asociado al username (limpiamos previos)
                cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (u,))
                cur.execute("INSERT INTO recuperacion_codigos (username, email, codigo, expiracion) VALUES (%s, %s, %s, %s)", (u, e, codigo, expiracion))
                db.commit()
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurri√≥ un error al generar el c√≥digo:\n{err}", parent=win)
                try:
                    if db.open:
                        cur.close(); db.close()
                except Exception:
                    pass
                return
            finally:
                try:
                    if db.open:
                        cur.close(); db.close()
                except Exception:
                    pass

            # Enviar el c√≥digo por correo
            if not enviar_email_cod(e, codigo, tipo='registro', username=u):
                return messagebox.showerror("Error de Correo", "No se pudo enviar el c√≥digo al correo proporcionado.", parent=win)

            # Mostrar formulario de confirmaci√≥n de c√≥digo
            confirm_win = ctk.CTkToplevel(win)
            confirm_win.title("üîí Confirmar creaci√≥n de usuario")
            confirm_win.minsize(380, 160)
            confirm_win.transient(win)
            confirm_win.after(10, confirm_win.grab_set)

            ctk.CTkLabel(confirm_win, text=f"Se ha enviado un c√≥digo al correo {e}. Ingresa el c√≥digo para confirmar la creaci√≥n del usuario.").pack(fill="x", padx=12, pady=(12,6))
            entry_code = ctk.CTkEntry(confirm_win, placeholder_text="C√≥digo de 6 d√≠gitos")
            entry_code.pack(fill="x", padx=12, pady=(0,8))

            def confirmar_registro():
                code = entry_code.get().strip()
                if not code:
                    return messagebox.showwarning("C√≥digo requerido", "Ingresa el c√≥digo enviado al correo.", parent=confirm_win)
                db2 = conectar_db()
                if not db2:
                    return
                try:
                    cur2 = db2.cursor()
                    cur2.execute("SELECT username FROM recuperacion_codigos WHERE username=%s AND email=%s AND codigo=%s AND expiracion > NOW()", (u, e, code))
                    res = cur2.fetchone()
                    if not res:
                        return messagebox.showwarning("C√≥digo inv√°lido", "C√≥digo inv√°lido o expirado.", parent=confirm_win)

                    # Registrar usuario
                    try:
                        h = hashlib.sha256(p.encode()).hexdigest()
                        db3 = conectar_db()
                        if not db3:
                            return
                        cur3 = db3.cursor()
                        cur3.execute("INSERT INTO usuarios_app (username, password_hash, email, es_admin) VALUES (%s, %s, %s, %s)", (u, h, e, val_admin.get()))
                        db3.commit()
                        vals_dict = {"username": u, "email": e, "es_admin": val_admin.get(), "password_hash": "***"}
                        registrar_historial(self.usuario, "INSERT", "usuarios_app", u, "", str(vals_dict))
                        messagebox.showinfo("√âxito", "Usuario creado correctamente.", parent=win)
                        # limpiar c√≥digo
                        cur2.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (u,))
                        db2.commit()
                        try:
                            if db3.open:
                                cur3.close(); db3.close()
                        except Exception:
                            pass
                        confirm_win.destroy()
                        win.destroy()
                    except IntegrityError:
                        messagebox.showerror("Error", "El usuario o correo ya existe.", parent=confirm_win)
                    except MySQLError as err:
                        messagebox.showerror("Error", f"No se pudo crear el usuario:\n{err}", parent=confirm_win)
                finally:
                    try:
                        if db2.open:
                            cur2.close(); db2.close()
                    except Exception:
                        pass

            ctk.CTkButton(confirm_win, text="Confirmar y crear usuario", command=confirmar_registro).pack(pady=10)
            entry_code.bind('<Return>', lambda ev: confirmar_registro())

        ctk.CTkButton(win, text="üíæ Guardar Usuario", command=guardar_user).pack(pady=20)
        _bind_responsive_wrap(win, [lbl_user, lbl_email, lbl_pass])

    def recuperar_contrasena(self):
        """Muestra la ventana para recuperar la contrase√±a del usuario."""
        win = ctk.CTkToplevel(self)
        win.title("üîë Recuperar Contrase√±a")
        win.minsize(360, 260)
        win.resizable(False, False)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        frame1 = ctk.CTkFrame(win)
        frame1.pack(fill="both", expand=True, padx=12, pady=12)
        lbl_ident = ctk.CTkLabel(frame1, text="Ingresa tu correo o nombre de usuario:")
        lbl_ident.pack(pady=(16, 6), anchor="w")
        entry = ctk.CTkEntry(frame1)
        entry.pack(fill="x")

        resend_btn = None
        resend_timer_label = None
        resend_seconds = [0]  # Usamos lista para mutabilidad en closure

        def enviar_codigo():
            """Env√≠a un c√≥digo de recuperaci√≥n por correo electr√≥nico."""
            identificador = entry.get().strip()
            if not identificador:
                return messagebox.showwarning("Campo requerido", "Por favor, ingresa tu usuario o correo.", parent=win)

            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                cur.execute("SELECT email, username FROM usuarios_app WHERE username=%s OR email=%s", (identificador, identificador))
                res = cur.fetchone()
                if not res:
                    return messagebox.showwarning("No encontrado", "Usuario o correo no encontrado.", parent=win)

                email = res[0]
                username = res[1]
                codigo = "".join(random.choices("0123456789", k=6))

                if enviar_email_cod(email, codigo, tipo='recuperacion', username=username):
                    expiracion = datetime.now(timezone.utc) + timedelta(minutes=15)
                    cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (username,))
                    cur.execute("INSERT INTO recuperacion_codigos (username, email, codigo, expiracion) VALUES (%s, %s, %s, %s)", (username, email, codigo, expiracion))
                    db.commit()
                    messagebox.showinfo("C√≥digo enviado", "C√≥digo enviado. Revisa tu correo.", parent=win)
                    frame1.pack_forget()
                    frame2.pack(fill="both", expand=True, padx=20, pady=20)
                    iniciar_resend_timer()
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurri√≥ un error:\n{err}", parent=win)
            finally:
                if db.open:
                    cur.close()
                    db.close()

        def iniciar_resend_timer():
            resend_seconds[0] = 180
            if resend_btn:
                resend_btn.configure(state="disabled")
            actualizar_timer()

        def actualizar_timer():
            if resend_seconds[0] > 0:
                if resend_timer_label:
                    resend_timer_label.configure(text=f"Puedes reenviar en {resend_seconds[0]} segundos")
                resend_seconds[0] -= 1
                win.after(1000, actualizar_timer)
            else:
                if resend_btn:
                    resend_btn.configure(state="normal")
                if resend_timer_label:
                    resend_timer_label.configure(text="Puedes volver a enviar el c√≥digo")

        ctk.CTkButton(frame1, text="Enviar C√≥digo", command=enviar_codigo).pack(pady=14)

        frame2 = ctk.CTkFrame(win)
        lbl_code = ctk.CTkLabel(frame2, text="Ingresa el c√≥digo:")
        lbl_code.pack(pady=(16, 6), anchor="w", padx=12)
        entry_codigo = ctk.CTkEntry(frame2, placeholder_text="C√≥digo de 6 d√≠gitos")
        entry_codigo.pack(fill="x", padx=12)
        lbl_newp = ctk.CTkLabel(frame2, text="Nueva Contrase√±a:")
        lbl_newp.pack(pady=(8, 4), anchor="w", padx=12)
        entry_new_pass_frame = PasswordEntry(frame2)
        entry_new_pass_frame.pack(fill="x", padx=12)
        # Adjuntar checklist de contrase√±a
        try:
            attach_password_checklist(frame2, entry_new_pass_frame)
        except Exception:
            pass

        def actualizar_contrasena():
            """Actualiza la contrase√±a del usuario si el c√≥digo es v√°lido."""
            codigo = entry_codigo.get().strip()
            new_pass = entry_new_pass_frame.get().strip()
            identificador = entry.get().strip()

            if not codigo or not new_pass:
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)

            validation_error = validar_contrasena(new_pass)
            if validation_error:
                return messagebox.showwarning("Contrase√±a d√©bil", validation_error, parent=win)

            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                cur.execute("SELECT username FROM recuperacion_codigos WHERE (username=%s OR email=%s) AND codigo=%s AND expiracion > NOW()", (identificador, identificador, codigo))
                res = cur.fetchone()

                if res:
                    h = hashlib.sha256(new_pass.encode()).hexdigest()
                    cur.execute("UPDATE usuarios_app SET password_hash=%s WHERE username=%s", (h, res[0]))
                    cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (res[0],))
                    db.commit()

                    registrar_historial("Sistema", "UPDATE", "usuarios_app", res[0], "password_hash: ****", "password_hash: ****")

                    messagebox.showinfo("√âxito", "Contrase√±a actualizada con √©xito.", parent=win)
                    win.destroy()
                else:
                    messagebox.showwarning("C√≥digo inv√°lido", "C√≥digo inv√°lido o expirado.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurri√≥ un error:\n{err}", parent=win)
            finally:
                if db.open:
                    cur.close()
                    db.close()

        btn_confirmar = ctk.CTkButton(frame2, text="Confirmar c√≥digo y actualizar contrase√±a", command=actualizar_contrasena)
        btn_confirmar.pack(pady=10, padx=12)

        resend_btn = ctk.CTkButton(frame2, text="Volver a enviar c√≥digo", command=lambda: [enviar_codigo(), iniciar_resend_timer()])
        resend_btn.pack(pady=(10, 0), padx=12)
        resend_timer_label = ctk.CTkLabel(frame2, text="")
        resend_timer_label.pack(pady=(4, 12), padx=12)
        _bind_responsive_wrap(win, [lbl_ident, lbl_code, lbl_newp])

    def cambiar_datos(self):
        """Muestra una ventana para que el usuario cambie su correo o contrase√±a."""
        win = ctk.CTkToplevel(self)
        win.title("‚öôÔ∏è Cambiar Mis Datos")
        win.minsize(420, 360)
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        db = conectar_db()
        if not db:
            return
        email_actual = ""
        try:
            cur = db.cursor()
            cur.execute("SELECT email FROM usuarios_app WHERE username=%s", (self.usuario,))
            result = cur.fetchone()
            if result:
                email_actual = result[0]
        except MySQLError:
            email_actual = ""
        finally:
            if db.open:
                cur.close()
                db.close()

        lbl_user = ctk.CTkLabel(win, text=f"Usuario: {self.usuario}", font=ctk.CTkFont(size=16, weight="bold"))
        lbl_user.pack(pady=(14, 8), anchor="w", padx=16)
        lbl_email = ctk.CTkLabel(win, text="Nuevo Correo Electr√≥nico: (dejar igual si no se quiere cambiar el correo electr√≥nico)")
        lbl_email.pack(pady=(8, 4), anchor="w", padx=16)
        e_email = ctk.CTkEntry(win)
        e_email.insert(0, email_actual)
        e_email.pack(padx=16, fill="x")
        lbl_pass = ctk.CTkLabel(win, text="Nueva Contrase√±a: (Deja en blanco para no cambiarla)")
        lbl_pass.pack(pady=(8, 4), anchor="w", padx=16)
        e_pass_frame = PasswordEntry(win)
        e_pass_frame.pack(padx=16, fill="x")
        # Adjuntar checklist de contrase√±a
        try:
            attach_password_checklist(win, e_pass_frame)
        except Exception:
            pass

        def actualizar_datos():
            """Inicia el flujo: env√≠a un c√≥digo al email actual (o al nuevo) y solicita confirmaci√≥n para aplicar cambios."""
            new_email = e_email.get().strip()
            new_pass = e_pass_frame.get().strip()

            if not validar_email(new_email):
                return messagebox.showwarning("Correo inv√°lido", "Formato de correo inv√°lido.", parent=win)

            # Elegir a qu√© correo enviar el c√≥digo: preferir el email actual si no cambia, sino al nuevo
            destino = email_actual if new_email == email_actual else new_email

            codigo = "".join(random.choices("0123456789", k=6))
            expiracion = datetime.now(timezone.utc) + timedelta(minutes=15)

            db = conectar_db()
            if not db:
                return
            try:
                cur = db.cursor()
                cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (self.usuario,))
                cur.execute("INSERT INTO recuperacion_codigos (username, email, codigo, expiracion) VALUES (%s, %s, %s, %s)", (self.usuario, destino, codigo, expiracion))
                db.commit()
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurri√≥ un error al generar el c√≥digo:\n{err}", parent=win)
                try:
                    if db.open: cur.close(); db.close()
                except Exception:
                    pass
                return
            finally:
                try:
                    if db.open: cur.close(); db.close()
                except Exception:
                    pass

            if not enviar_email_cod(destino, codigo, tipo='cambio', username=self.usuario, email_destino=new_email):
                return messagebox.showerror("Error de Correo", "No se pudo enviar el c√≥digo al correo destino.", parent=win)

            # Pedir confirmaci√≥n del c√≥digo
            confirm_win = ctk.CTkToplevel(win)
            confirm_win.title("üîí Confirmar cambios")
            confirm_win.minsize(380, 160)
            confirm_win.transient(win)
            confirm_win.after(10, confirm_win.grab_set)

            ctk.CTkLabel(confirm_win, text=f"Se ha enviado un c√≥digo al correo {destino}. Ingresa el c√≥digo para confirmar los cambios.").pack(fill="x", padx=12, pady=(12,6))
            entry_code = ctk.CTkEntry(confirm_win, placeholder_text="C√≥digo de 6 d√≠gitos")
            entry_code.pack(fill="x", padx=12, pady=(0,8))

            def confirmar_cambios():
                code = entry_code.get().strip()
                if not code:
                    return messagebox.showwarning("C√≥digo requerido", "Ingresa el c√≥digo enviado al correo.", parent=confirm_win)
                db2 = conectar_db()
                if not db2:
                    return
                try:
                    cur2 = db2.cursor()
                    cur2.execute("SELECT username FROM recuperacion_codigos WHERE username=%s AND email=%s AND codigo=%s AND expiracion > NOW()", (self.usuario, destino, code))
                    res = cur2.fetchone()
                    if not res:
                        return messagebox.showwarning("C√≥digo inv√°lido", "C√≥digo inv√°lido o expirado.", parent=confirm_win)

                    # Aplicar cambios
                    try:
                        db3 = conectar_db()
                        if not db3:
                            return
                        cur3 = db3.cursor()
                        if new_pass:
                            validation_error = validar_contrasena(new_pass)
                            if validation_error:
                                return messagebox.showwarning("Contrase√±a d√©bil", validation_error, parent=confirm_win)
                            h = hashlib.sha256(new_pass.encode()).hexdigest()
                            cur3.execute("UPDATE usuarios_app SET email=%s, password_hash=%s WHERE username=%s", (new_email, h, self.usuario))
                            registrar_historial(self.usuario, "UPDATE", "usuarios_app", self.usuario, f"email: {email_actual}, password:***", f"email: {new_email}, password:***")
                        else:
                            cur3.execute("UPDATE usuarios_app SET email=%s WHERE username=%s", (new_email, self.usuario))
                            registrar_historial(self.usuario, "UPDATE", "usuarios_app", self.usuario, f"email: {email_actual}", f"email: {new_email}")
                        db3.commit()
                        # limpiar codigo
                        cur2.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (self.usuario,))
                        db2.commit()
                        messagebox.showinfo("√âxito", "Datos actualizados correctamente.", parent=win)
                        try:
                            if db3.open: cur3.close(); db3.close()
                        except Exception:
                            pass
                        confirm_win.destroy()
                        win.destroy()
                    except IntegrityError:
                        messagebox.showerror("Error", "El correo ya est√° en uso.", parent=confirm_win)
                    except MySQLError as err:
                        messagebox.showerror("Error", f"No se pudieron actualizar los datos:\n{err}", parent=confirm_win)
                finally:
                    try:
                        if db2.open: cur2.close(); db2.close()
                    except Exception:
                        pass

            ctk.CTkButton(confirm_win, text="Confirmar cambios", command=confirmar_cambios).pack(pady=10)
            entry_code.bind('<Return>', lambda ev: confirmar_cambios())

        ctk.CTkButton(win, text="üíæ Guardar Cambios", command=actualizar_datos).pack(pady=16)
        _bind_responsive_wrap(win, [lbl_user, lbl_email, lbl_pass])

    def cargar_tabla(self, tabla):
        """
        Carga los datos de la tabla seleccionada en el Treeview.
        """
        # Si no hay tablas, limpia el Treeview
        if tabla == "No hay tablas":
            self.tree.delete(*self.tree.get_children())
            self.tree["columns"] = []
            return
        # Muestra el overlay animado de carga
        try:
            self._start_loading(parent=self)
        except Exception:
            pass

        self.update_idletasks()

        # Elimina los datos actuales de la tabla
        self.tree.delete(*self.tree.get_children())
        try:
            # Obtiene columnas y filas de la tabla seleccionada
            self.col, rows = obtener_datos(tabla)
        finally:
            try:
                self._stop_loading()
            except Exception:
                pass
        # Si no hay columnas, termina
        if not self.col:
            return

        # Configura el Treeview con las nuevas columnas y datos
        self.table = tabla
        self.tree["columns"] = self.col
        self.tree["show"] = "headings"
        for c in self.col:
            self.tree.heading(c, text=c)
            self.tree.column(c, anchor="w", width=150, minwidth=100)
        for r in rows:
            self.tree.insert("", "end", values=r)

    def cargar_tabla_actual(self):
        """Recarga la tabla que se est√° mostrando actualmente."""
        if hasattr(self, "table"):
            self.cargar_tabla(self.table)

    def buscar(self, ev=None):
        """Filtra los datos de la tabla seg√∫n el texto de b√∫squeda."""
        if not hasattr(self, "table"): return
        txt = self.busc.get().lower()
        self.tree.delete(*self.tree.get_children())
        _, rows = obtener_datos(self.table)
        for r in rows:
            if any(txt in str(v).lower() for v in r):
                self.tree.insert("", "end", values=r)

    def abrir_form(self):
        """Abre un formulario para a√±adir un nuevo registro a la tabla actual."""
        if not hasattr(self, "table"):
            return mostrar_toast("Selecciona una tabla primero", self)
        win = ctk.CTkToplevel(self)
        win.title(f"‚ûï Agregar registro a {self.table}")
        win.minsize(420, 420)
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        scroll = ctk.CTkScrollableFrame(win)
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        cols, _ = obtener_datos(self.table)
        cols_no_id = [c for c in cols if c.lower() != "id"]
        entries = {}

        for c in cols_no_id:
            lbl = ctk.CTkLabel(scroll, text=f"{c}:")
            lbl.pack(anchor="w", padx=6, pady=(8, 0))
            if "fecha" in c.lower():
                ent = ctk.CTkEntry(scroll)
                ent.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                ent.configure(state="readonly")
                ent.pack(fill="x", padx=5)
                entries[c] = ent
            elif "pass" in c.lower() and "password_hash" not in c.lower():
                ent = PasswordEntry(scroll)
                ent.pack(fill="x", padx=6)
                entries[c] = ent
                # Adjuntar checklist para este campo de contrase√±a
                try:
                    attach_password_checklist(scroll, ent)
                except Exception:
                    pass
            else:
                ent = ctk.CTkEntry(scroll)
                ent.pack(fill="x", padx=6)
                entries[c] = ent

        def insertar():
            """Inserta el nuevo registro en la base de datos."""
            vals_dict = {}
            for c in cols_no_id:
                v = entries[c].get()
                if "pass" in c.lower() and "password_hash" not in c.lower():
                    validation_error = validar_contrasena(v)
                    if validation_error:
                        return messagebox.showwarning("Contrase√±a d√©bil", validation_error, parent=win)
                    v = hashlib.sha256(v.encode()).hexdigest()
                vals_dict[c] = v

            vals = list(vals_dict.values())

            if any(not str(x).strip() for x in vals):
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)

            db = conectar_db()
            if not db: return

            try:
                cur = db.cursor()
                # Se construye la consulta de forma segura
                query = f"INSERT INTO `{self.table}` ({','.join([f'`{c}`' for c in cols_no_id])}) VALUES ({','.join(['%s'] * len(vals))})"
                cur.execute(query, vals)
                rid = cur.lastrowid
                db.commit()

                if "password_hash" in vals_dict:
                    vals_dict["password_hash"] = "***"

                registrar_historial(self.usuario, "INSERT", self.table, rid, "", str(vals_dict))
                win.destroy()
                self.cargar_tabla_actual()
                messagebox.showinfo("√âxito", "Registro insertado.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error al Insertar", f"No se pudo guardar:\n{err}", parent=win)
            finally:
                if db.open:
                    cur.close()
                    db.close()

        ctk.CTkButton(scroll, text="üíæ Insertar Registro", command=insertar).pack(pady=16, padx=6)

    def eliminar_registro(self):
        """
        Elimina un registro seleccionado del Treeview y la base de datos.
        """
        # Obtiene el registro seleccionado
        sel = self.tree.focus()
        if not sel: return

        # Extrae los valores y la clave primaria
        r = self.tree.item(sel)["values"]
        pk_col_name = self.col[0]
        pk_value = r[0]
        before = str(dict(zip(self.col, r)))

        # Pide confirmaci√≥n al usuario antes de borrar
        if messagebox.askyesno("Confirmar Borrado", f"¬øSeguro que quieres eliminar el registro con {pk_col_name} = {pk_value}?"):
            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                # Ejecuta el borrado seguro
                cur.execute(f"DELETE FROM `{self.table}` WHERE `{pk_col_name}`=%s", (pk_value,))
                db.commit()
                # Registra el borrado en el historial
                registrar_historial(self.usuario, "DELETE", self.table, pk_value, before, "")
                # Recarga la tabla y muestra mensaje de √©xito
                self.cargar_tabla_actual()
                messagebox.showinfo("√âxito", "Registro eliminado.", parent=self)
            except IntegrityError as e:
                # Manejo de error de clave for√°nea (no se puede borrar si hay dependencias)
                if e.args[0] == 1451:
                    match = re.search(r"fails \(`\w+`\.`(.+?)`\)", e.args[1])
                    tabla_dependiente = match.group(1) if match else "otra tabla"
                    msg = (f"No se puede eliminar. Este registro est√° siendo usado por la tabla '{tabla_dependiente}'.\n\n"
                           f"Debes eliminar o actualizar los registros dependientes primero.")
                    messagebox.showerror("Error de Integridad", msg)
                else:
                    messagebox.showerror("Error de BD", f"No se pudo eliminar:\n{e}")
            except MySQLError as err:
                messagebox.showerror("Error de Borrado", f"Ocurri√≥ un error:\n{err}")
            finally:
                if db.open: cur.close(); db.close()

    def editar_celda(self, ev):
        """
        Permite editar el valor de una celda directamente en el Treeview.
        """
        # Solo permite editar si el usuario es administrador
        if not self.is_admin: return

        # Identifica la fila y columna seleccionada
        sel = self.tree.identify_row(ev.y)
        cid = self.tree.identify_column(ev.x)
        if not sel or not cid: return

        # Obtiene el √≠ndice y nombre de la columna
        col_index = int(cid.replace("#", "")) - 1
        col_name = self.col[col_index]
        pk_col_name = self.col[0]
        pk_value = self.tree.item(sel)["values"][0]

        # No permite editar la clave primaria
        if col_name.lower() == pk_col_name.lower():
            return mostrar_toast(f"La clave primaria ({pk_col_name}) no se puede editar.", self)

        old_val = self.tree.item(sel)["values"][col_index]

        # Obtiene la posici√≥n y tama√±o de la celda
        x, y, w, h = self.tree.bbox(sel, cid)

        # Determina si el campo es de tipo contrase√±a
        is_password_field = "pass" in col_name.lower() and "hash" not in col_name.lower()

        # Crea el widget de edici√≥n adecuado
        if is_password_field:
            ent = PasswordEntry(self.tree, width=w, height=h)
            ent.place(x=x, y=y)
            ent.entry.insert(0, old_val)
        else:
            ent = ctk.CTkEntry(self.tree, width=w, height=h)
            ent.place(x=x, y=y)
            ent.insert(0, old_val)

        ent.focus()

        def save(event):
            """
            Guarda el nuevo valor de la celda en la base de datos.
            """
            try:
                new_val = ent.get()
                ent.destroy()
            except TclError:
                return 

            # Si no hay cambios, termina
            if new_val == old_val: return

            final_val = new_val
            # Si es campo de contrase√±a, valida y hashea
            if is_password_field:
                validation_error = validar_contrasena(new_val)
                if validation_error:
                    return messagebox.showwarning("Contrase√±a d√©bil", validation_error, parent=self)
                final_val = hashlib.sha256(new_val.encode()).hexdigest()

            db = conectar_db()
            if not db: return

            try:
                cur = db.cursor()
                # Actualiza el valor en la base de datos
                cur.execute(f"UPDATE `{self.table}` SET `{col_name}`=%s WHERE `{pk_col_name}`=%s", (final_val, pk_value))
                db.commit()

                # Registra el cambio en el historial
                antes = f"{col_name}: {old_val}"
                despues = f"{col_name}: {final_val}"
                if "pass" in col_name.lower():
                    antes = f"{col_name}: ***"
                    despues = f"{col_name}: ***"

                registrar_historial(self.usuario, "UPDATE", self.table, pk_value, antes, despues)
                self.cargar_tabla_actual()

                # Muestra notificaci√≥n de √©xito
                mostrar_toast("Campo actualizado.", self)
            except MySQLError as err:
                messagebox.showerror("Error al Actualizar", f"No se pudo guardar:\n{err}")
                self.cargar_tabla_actual()
            finally:
                if db.open: cur.close(); db.close()

        ent.bind("<Return>", save)
        ent.bind("<FocusOut>", lambda e: ent.destroy())

    def ver_historial(self):
        """Muestra una ventana con el historial completo de cambios en la base de datos."""
        win = ctk.CTkToplevel(self)
        win.title("üìã Historial de Cambios")
        # Hacemos la ventana un poco m√°s peque√±a por defecto y permitimos redimensionar
        win.minsize(600, 300)
        win.geometry("700x360")
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=8, pady=8)

        style = ttk.Style(frame)
        style.theme_use("default")
        style.configure("Treeview", background="#2a2d2e", foreground="white", rowheight=25, fieldbackground="#343638", borderwidth=0)
        style.map('Treeview', background=[('selected', '#245d81')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat", font=('Calibri', 10, 'bold'))
        style.map("Treeview.Heading", background=[('active', '#343638')])

        tv = ttk.Treeview(frame, style="Treeview")
        tv["columns"] = ("ID", "Usuario", "Fecha", "Acci√≥n", "Tabla", "ID Reg.", "Valores Anteriores", "Valores Nuevos")
        tv.column("#0", width=0, stretch="no")

        tv.column("ID", width=50, minwidth=50, stretch="no")
        tv.column("Usuario", width=120, minwidth=100, stretch="no")
        tv.column("Fecha", width=150, minwidth=130, stretch="no")
        tv.column("Acci√≥n", width=100, minwidth=80, stretch="no")
        tv.column("Tabla", width=120, minwidth=100, stretch="no")
        tv.column("ID Reg.", width=80, minwidth=70, stretch="no")
        tv.column("Valores Anteriores", anchor="w", stretch="yes", minwidth=400, width=500)
        tv.column("Valores Nuevos", anchor="w", stretch="yes", minwidth=400, width=500)

        for col in tv["columns"]:
            tv.heading(col, text=col, anchor="w")

        db = conectar_db()
        if not db:
            return

        try:
            cur = db.cursor()
            # Si el usuario no es admin, solo mostrar sus propias acciones y limitar a 200 filas
            if self.is_admin:
                cur.execute("SELECT id, usuario, fecha, accion, tabla, registro_id, valores_antes, valores_despues FROM historial ORDER BY fecha DESC LIMIT 2000")
            else:
                cur.execute("SELECT id, usuario, fecha, accion, tabla, registro_id, valores_antes, valores_despues FROM historial WHERE usuario=%s ORDER BY fecha DESC LIMIT 200", (self.usuario,))
            for r in cur.fetchall():
                tv.insert("", "end", values=r)
        except MySQLError as err:
            messagebox.showerror("Error", f"No se pudo cargar el historial:\n{err}", parent=win)
        finally:
            try:
                if db.open:
                    cur.close()
                    db.close()
            except Exception:
                pass

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tv.xview)
        tv.configure(xscrollcommand=hsb.set)

        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

    def mostrar_ayuda(self):
        """Muestra una ventana de ayuda y preguntas frecuentes."""
        # Si ya hay un dropdown abierto, ci√©rralo (toggle)
        try:
            if getattr(self, 'help_dropdown', None) and self.help_dropdown.winfo_exists():
                try:
                    self.help_dropdown.destroy()
                except Exception:
                    pass
                self.help_dropdown = None
                return
        except Exception:
            pass

        # crear un Toplevel sin decoraciones para que act√∫e como un "dropdown" centrado en el bot√≥n
        win = ctk.CTkToplevel(self)
        self.help_dropdown = win
        # quitar bordes/decoration para apariencia de men√∫
        try:
            win.overrideredirect(True)
        except Exception:
            # en algunos entornos customtkinter puede no soportarlo; ignorar
            pass
        # dar una clase de t√≠tulo interna (√∫til para debug)
        win.title("Ayuda")

        # tama√±o razonable; el contenido puede expandirse
        DEFAULT_W = 460
        DEFAULT_H = 340
        win.geometry(f"{DEFAULT_W}x{DEFAULT_H}")

        # posicionar centrado respecto al bot√≥n
        try:
            btn = getattr(self, 'btn_help', None)
            win.update_idletasks()
            w = DEFAULT_W
            h = DEFAULT_H
            screen_w = win.winfo_screenwidth()
            screen_h = win.winfo_screenheight()

            if btn:
                bx = btn.winfo_rootx()
                by = btn.winfo_rooty()
                bw = btn.winfo_width()
                bh = btn.winfo_height()

                x = bx + (bw - w) // 2
                y = by + bh + 6

                # ajustar para que no quede fuera de pantalla
                if x + w + 8 > screen_w:
                    x = max(8, bx + bw - w)
                if x < 8:
                    x = 8

                # si no cabe por abajo, mostrar por arriba del bot√≥n
                if y + h + 8 > screen_h:
                    y = max(8, by - h - 6)

                win.geometry(f"+{int(x)}+{int(y)}")
            else:
                # centrado en pantalla si no hay bot√≥n
                x = max(8, (screen_w - w)//2)
                y = max(8, (screen_h - h)//4)
                win.geometry(f"+{int(x)}+{int(y)}")
        except Exception:
            pass

        # cerrar al perder foco o al pulsar Escape
        def _on_focus_out(event=None):
            try:
                if self.help_dropdown and self.help_dropdown.winfo_exists():
                    self.help_dropdown.destroy()
            except Exception:
                pass
            self.help_dropdown = None

        win.bind("<FocusOut>", _on_focus_out)
        win.bind("<Escape>", lambda e: _on_focus_out())

        # forzar foco al dropdown para poder detectat FocusOut cuando el usuario haga clic fuera
        try:
            win.focus_force()
        except Exception:
            pass
        scroll_frame = ctk.CTkScrollableFrame(win)
        scroll_frame.pack(fill="both", expand=True, padx=8, pady=8)

        # El contenido de la ayuda var√≠a si el usuario es administrador
        if self.is_admin:
            ctk.CTkLabel(scroll_frame, text="Ayuda para Administradores", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(10, 5))
            faq = {
                "¬øC√≥mo a√±ado un nuevo registro?": "Selecciona una tabla y haz clic en '‚ûï A√±adir'. Completa el formulario y haz clic en 'Guardar'.",
                "¬øC√≥mo edito un dato?": "Haz doble clic sobre la celda que deseas modificar. Escribe el nuevo valor y presiona 'Enter'.",
                "¬øPor qu√© no puedo eliminar un registro?": "Es posible que el registro est√© siendo usado como clave for√°nea en otra tabla. Debes eliminar primero los registros dependientes.",
                "¬øC√≥mo creo un nuevo usuario o cambio sus datos?": "Haz clic en 'üë§ Nuevos usuarios' para crear una nueva cuenta. Puedes usar '‚öôÔ∏è Cambiar mis datos' para modificar tu informaci√≥n.",
                "¬øQu√© informaci√≥n me muestra el historial?": "El 'Historial de Cambios' registra todas las acciones de inserci√≥n, actualizaci√≥n y eliminaci√≥n de datos, incluyendo qui√©n las realiz√≥ y cu√°ndo. Los datos sensibles como contrase√±as y claves de usuarios est√°n enmascarados para mayor seguridad.",
                "Mi contrase√±a no es aceptada, ¬øpor qu√©?": "Las contrase√±as deben tener un m√≠nimo de 8 caracteres, al menos una letra may√∫scula y un car√°cter especial.",
                "Olvid√© mi contrase√±a, ¬øqu√© hago?": "En la ventana de login, haz clic en 'Olvid√© mi contrase√±a' y sigue las instrucciones para recibir un c√≥digo de recuperaci√≥n por correo.",
                "No puedo conectar a la base de datos": "Verifica que el servidor MySQL est√© corriendo y que la configuraci√≥n en tu archivo '.env' sea correcta (host, usuario, contrase√±a, etc.)."
            }
        else:
            ctk.CTkLabel(scroll_frame, text="Ayuda para Usuarios", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(10, 5))
            faq = {
                "¬øC√≥mo a√±ado un nuevo registro?": "Selecciona una tabla y haz clic en '‚ûï A√±adir'. Completa el formulario y haz clic en 'Guardar'.",
                "¬øC√≥mo edito un dato?": "Haz doble clic sobre la celda que deseas modificar. Escribe el nuevo valor y presiona 'Enter'.",
                "¬øPor qu√© no puedo eliminar un registro?": "Tu nivel de usuario no tiene permisos para borrar registros. Contacta a un administrador para realizar esta acci√≥n.",
                "Mi contrase√±a no es aceptada, ¬øpor qu√©?": "Las contrase√±as deben tener un m√≠nimo de 8 caracteres, al menos una letra may√∫scula y un car√°cter especial.",
                "Olvid√© mi contrase√±a, ¬øqu√© hago?": "En la ventana de login, haz clic en 'Olvid√© mi contrase√±a' y sigue las instrucciones para recibir un c√≥digo de recuperaci√≥n por correo."
            }

        lbls = []
        for pregunta, respuesta in faq.items():
            q = ctk.CTkLabel(scroll_frame, text=f"**{pregunta}**", font=ctk.CTkFont(size=14, weight="bold"), justify="left")
            q.pack(anchor="w", pady=(10, 0))
            a = ctk.CTkLabel(scroll_frame, text=respuesta, font=ctk.CTkFont(size=12), justify="left")
            a.pack(anchor="w", pady=(0, 6))
            lbls.extend([q, a])
        _bind_responsive_wrap(win, lbls)

# -----------------------------------------------
# PUNTO DE ENTRADA DE LA APLICACI√ìN
# -----------------------------------------------
if __name__ == "__main__":
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    app = App()
    app.mainloop()