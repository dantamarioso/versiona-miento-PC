# =================================================================================
# APLICACIÓN DE GESTIÓN DE BASE DE DATOS "N.I.C.O.L.E" - VERSIÓN MEJORADA
# =================================================================================

# -------------------------
# IMPORTACIÓN DE LIBRERÍAS
# -------------------------
# Importa todas las bibliotecas necesarias para la interfaz de usuario (customtkinter),
# la base de datos (pymysql), seguridad (hashlib), y otras funcionalidades.
# Se usa dotenv para cargar las credenciales desde un archivo .env.

import hashlib
import re
import smtplib
import ssl
import random
from datetime import datetime, timedelta
from tkinter import messagebox, ttk, TclError, filedialog

import customtkinter as ctk
import pymysql
from PIL import Image, ImageTk
from pymysql.err import MySQLError, IntegrityError

import csv
import openpyxl
from fpdf import FPDF
from dotenv import load_dotenv

import os

# Carga las variables de entorno desde el archivo .env
load_dotenv()

# --------------------------------
# CONFIGURACIÓN DE CREDENCIALES
# --------------------------------
# Carga las credenciales del servidor de correo electrónico desde el archivo .env
EMAIL_HOST = os.getenv("EMAIL_HOST")
EMAIL_PORT = os.getenv("EMAIL_PORT")
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")

# -----------------------------------------------
# FUNCIONES DE VALIDACIÓN Y GESTIÓN DE LA BASE DE DATOS
# -----------------------------------------------

def conectar_db():
    """Establece y retorna una conexión con la base de datos MySQL."""
    try:
        return pymysql.connect(
            host=os.getenv("DB_HOST"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            database=os.getenv("DB_DATABASE"),
            port=int(os.getenv("DB_PORT")),
            ssl={"ca": "isrgrootx1.pem"}  
        )
    except MySQLError as err:
        print(f"Error MySQL: No se pudo conectar a MySQL:\n{err}")
        return None

def init_db():
    """
    Inicializa la base de datos.
    Crea las tablas 'historial', 'usuarios_app' y 'recuperacion_codigos' si no existen.
    También crea un usuario 'admin' por defecto si no hay ninguno.
    """
    db = conectar_db()
    if not db:
        return

    try:
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historial (
                id INT AUTO_INCREMENT PRIMARY KEY,
                usuario VARCHAR(50),
                fecha DATETIME,
                accion VARCHAR(20),
                tabla VARCHAR(50),
                registro_id VARCHAR(50),
                valores_antes TEXT,
                valores_despues TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios_app (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50) UNIQUE,
                password_hash VARCHAR(255),
                email VARCHAR(255) UNIQUE,
                es_admin BOOLEAN
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS recuperacion_codigos (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50),
                email VARCHAR(255),
                codigo VARCHAR(6),
                expiracion DATETIME
            )
        """)

        # Crea el usuario admin si no existe
        cur.execute("SELECT * FROM usuarios_app WHERE username='admin'")
        if not cur.fetchone():
            h = hashlib.sha256("admin123".encode()).hexdigest()
            cur.execute("INSERT INTO usuarios_app (username, password_hash, email, es_admin) VALUES (%s, %s, %s, %s)", ("admin", h, "admin@example.com", True))

        db.commit()
    except MySQLError as err:
        messagebox.showerror("Error de Inicialización", f"No se pudo inicializar la base de datos:\n{err}")
    finally:
        if db.open:
            cur.close()
            db.close()

def validar_email(email):
    """Verifica si una cadena tiene un formato de correo electrónico válido."""
    regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.fullmatch(regex, email)

def validar_contrasena(password):
    """
    Verifica si una contraseña cumple con los requisitos de seguridad:
    - Mínimo 8 caracteres
    - Al menos una letra mayúscula
    - Al menos un carácter especial
    """
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres."
    if not any(char.isupper() for char in password):
        return "La contraseña debe contener al menos una letra mayúscula."
    if not re.search(r"[!@#$%^&*()_+={}\[\]|\\:;\"'<>,.?/`~]", password):
        return "La contraseña debe contener al menos un carácter especial."
    return None # La contraseña es válida

def enviar_email_cod(destinatario, codigo):
    """Envía un correo electrónico con un código de verificación."""
    if not (EMAIL_HOST and EMAIL_PORT and EMAIL_USER and EMAIL_PASS and EMAIL_SENDER):
        messagebox.showwarning("Configuración de Correo", "La configuración de correo no está completa en el archivo .env.")
        return False
    
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(EMAIL_HOST, int(EMAIL_PORT), context=context) as server:
            server.login(EMAIL_USER, EMAIL_PASS.replace(" ", ""))
            
            asunto = "Código de Recuperación de Contraseña N.I.C.O.L.E"
            cuerpo = f"""
            Hola,

            Has solicitado un código para recuperar tu contraseña.
            Tu código de verificación es: {codigo}

            Este código es válido por 15 minutos. Si no solicitaste este cambio,
            ignora este mensaje.

            Saludos,
            El equipo de N.I.C.O.L.E
            """
            
            msg = f"Subject: {asunto}\n\n{cuerpo}"
            server.sendmail(EMAIL_SENDER, destinatario, msg.encode('utf-8'))
        return True
    except Exception as e:
        messagebox.showerror("Error de Correo", f"No se pudo enviar el correo:\n{e}")
        return False

def registrar_historial(usuario, accion, tabla, registro_id, antes="", despues=""):
    """
    Registra una acción (INSERT, UPDATE, DELETE) en la tabla 'historial'.
    Enmascara datos sensibles como contraseñas para mayor seguridad.
    """
    db = conectar_db()
    if not db:
        return

    try:
        cur = db.cursor()
        # Enmascarar el ID del registro y la contraseña si la tabla es 'usuarios_app'
        if tabla == "usuarios_app":
            registro_id_hist = "***"
            
            # Enmascara el campo 'usuario' si está presente
            if "usuario" in despues:
                despues = re.sub(r"'usuario': '.*?'", "'usuario': '***'", despues)
            if "usuario" in antes:
                antes = re.sub(r"'usuario': '.*?'", "'usuario': '***'", antes)

            # Enmascara el campo 'correo' si está presente
            if "correo" in despues:
                despues = re.sub(r"'correo': '.*?'", "'correo': '***'", despues)
            if "correo" in antes:
                antes = re.sub(r"'correo': '.*?'", "'correo': '***'", antes)

            # Enmascara el campo 'password_hash' como ya lo tenías
            if "password_hash" in despues:
                despues = re.sub(r"'password_hash': '.*?'", "'password_hash': '***'", despues)
            if "password_hash" in antes:
                antes = re.sub(r"'password_hash': '.*?'", "'password_hash': '***'", antes)
        else:
            registro_id_hist = str(registro_id)

        cur.execute("""
            INSERT INTO historial (usuario, fecha, accion, tabla, registro_id, valores_antes, valores_despues)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (usuario, datetime.now(), accion, tabla, registro_id_hist, antes, despues))
        db.commit()
    except MySQLError as err:
        print(f"Error al registrar en historial: {err}")
    finally:
        if db.open:
            cur.close()
            db.close()

def obtener_tablas():
    """Obtiene una lista de las tablas de la BD, excluyendo las internas de la app."""
    db = conectar_db()
    if not db:
        return []

    try:
        cur = db.cursor()
        cur.execute("SHOW TABLES")
        tablas = [r[0] for r in cur.fetchall() if r[0] not in ("historial", "usuarios_app", "recuperacion_codigos")]
        return tablas
    except MySQLError as err:
        messagebox.showerror("Error de Lectura", f"No se pudieron obtener las tablas:\n{err}")
        return []
    finally:
        if db.open:
            cur.close()
            db.close()

def _validar_nombre_tabla(tabla, lista_tablas_permitidas):
    """Valida que un nombre de tabla esté en la lista de tablas permitidas."""
    if tabla in lista_tablas_permitidas:
        return True
    return False

def obtener_datos(tabla):
    """
    Obtiene las columnas y todas las filas de una tabla específica.
    Incluye una capa de seguridad para evitar la inyección de SQL.
    """
    db = conectar_db()
    if not db:
        return [], []

    tablas_validas = obtener_tablas()
    if not _validar_nombre_tabla(tabla, tablas_validas):
        messagebox.showerror("Error de Seguridad", "Nombre de tabla inválido.")
        return [], []

    try:
        cur = db.cursor()
        # Se construye la consulta de forma segura
        cur.execute(f"SELECT * FROM `{tabla}`")
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        return cols, rows
    except MySQLError as err:
        messagebox.showerror("Error de Lectura", f"No se pudieron obtener los datos de la tabla '{tabla}':\n{err}")
        return [], []
    finally:
        if db.open:
            cur.close()
            db.close()

def mostrar_toast(texto, parent):
    """Muestra una notificación emergente temporal (toast) en la pantalla."""
    toast = ctk.CTkToplevel(parent)
    toast.overrideredirect(True)
    
    x = parent.winfo_rootx() + parent.winfo_width() - 310
    y = parent.winfo_rooty() + parent.winfo_height() - 110
    toast.geometry(f"300x50+{x}+{y}")
    
    ctk.CTkLabel(toast, text=texto, font=ctk.CTkFont(size=14)).pack(expand=True, fill="both", padx=10, pady=5)
    toast.attributes("-topmost", True)
    toast.after(2500, toast.destroy)

# -----------------------------------------------
# CLASES DE LA INTERFAZ DE USUARIO
# -----------------------------------------------

class PasswordEntry(ctk.CTkFrame):
    """
    Widget de entrada de contraseña con un botón para mostrar/ocultar la contraseña.
    Es una mejora sobre el widget de entrada estándar de customtkinter.
    """
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)

        self.entry = ctk.CTkEntry(self, show="*")
        self.entry.grid(row=0, column=0, sticky="ew")

        try:
            self.eye_open_image = ctk.CTkImage(Image.open("eye_open.png"), size=(20, 20))
            self.eye_closed_image = ctk.CTkImage(Image.open("eye_closed.png"), size=(20, 20))
        except FileNotFoundError:
            self.eye_open_image = None
            self.eye_closed_image = None
            messagebox.showwarning("Icono no encontrado", "Asegúrate de tener 'eye_open.png' y 'eye_closed.png' en la carpeta.")

        self.show_pass_button = ctk.CTkButton(
            self,
            image=self.eye_open_image,
            text="",
            width=30,
            command=self.toggle_show_password,
            fg_color="transparent",
            hover_color=ctk.CTkEntry(self).cget("fg_color")
        )
        self.show_pass_button.grid(row=0, column=1, padx=(5, 0))

        self.is_password_visible = False

    def toggle_show_password(self):
        """Alterna la visibilidad de la contraseña."""
        if self.is_password_visible:
            self.entry.configure(show="*")
            self.show_pass_button.configure(image=self.eye_open_image)
        else:
            self.entry.configure(show="")
            self.show_pass_button.configure(image=self.eye_closed_image)
        self.is_password_visible = not self.is_password_visible

    def get(self):
        """Retorna el texto del campo de entrada."""
        return self.entry.get()
    
    def insert(self, index, string):
        """Inserta texto en el campo de entrada."""
        self.entry.insert(index, string)

    def pack(self, **kwargs):
        super().pack(**kwargs)
        self.entry.pack_propagate(False)

    def bind(self, sequence, func, add='+'):
        self.entry.bind(sequence, func, add=add)

class App(ctk.CTk):
    """
    Clase principal que encapsula toda la aplicación, manejando la UI y eventos.
    Contiene la lógica para el login, la visualización de tablas y la interacción con la base de datos.
    """
    def __init__(self):
        super().__init__()
        self.withdraw()
        init_db()
        self.usuario = None
        self.is_admin = False
        self.login_usuario()

    def login_usuario(self):
        """Muestra la ventana de login al iniciar la aplicación."""
        win = ctk.CTkToplevel(self)
        win.title("🔐 Login")
        win.geometry("500x350")
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", self.destroy)
        win.transient(self)
        win.after(100, win.grab_set)

        ctk.CTkLabel(win, text="Usuario:").pack(pady=(20, 5))
        entry_user = ctk.CTkEntry(win, placeholder_text="nombre de usuario")
        entry_user.pack(padx=20, fill="x")
        ctk.CTkLabel(win, text="Contraseña:").pack(pady=5)
        
        entry_pass_frame = PasswordEntry(win)
        entry_pass_frame.pack(padx=20, fill="x")

        def intentar(event=None):
            u = entry_user.get()
            p = entry_pass_frame.get()
            if not u or not p:
                mostrar_toast("Usuario y contraseña requeridos", win)
                return

            h = hashlib.sha256(p.encode()).hexdigest()
            db = conectar_db()
            if not db: return

            try:
                cur = db.cursor()
                cur.execute("SELECT es_admin FROM usuarios_app WHERE username=%s AND password_hash=%s", (u, h))
                res = cur.fetchone()
                if res:
                    self.usuario = u
                    self.is_admin = bool(res[0])
                    win.destroy()
                    self.mostrar_bienvenida()
                else:
                    mostrar_toast("Credenciales inválidas", win)
            except MySQLError as err:
                messagebox.showerror("Error de Login", f"Ocurrió un error:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()
        
        btn_frame = ctk.CTkFrame(win, fg_color="transparent")
        btn_frame.pack(pady=20, padx=20, fill="x")
        ctk.CTkButton(btn_frame, text="Entrar", command=intentar).pack(side="left", expand=True)
        ctk.CTkButton(btn_frame, text="Olvidé mi contraseña", command=self.recuperar_contrasena, fg_color="gray", hover_color="darkgray").pack(side="right", expand=True)

        entry_user.bind('<Return>', intentar)
        entry_pass_frame.bind('<Return>', intentar)

    def mostrar_bienvenida(self):
        """Muestra una ventana de bienvenida después de un login exitoso."""
        self.withdraw()
        win = ctk.CTkToplevel(self)
        win.title("🎉 Bienvenida")
        win.geometry("400x200")
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", self.destroy)
        win.transient(self)
        win.after(10, win.grab_set)

        ctk.CTkLabel(win, text=f"¡Bienvenido, {self.usuario}!", font=ctk.CTkFont(size=24, weight="bold")).pack(pady=(30, 10))
        ctk.CTkLabel(win, text="Has iniciado sesión correctamente.", justify="center").pack()

        def continuar():
            win.destroy()
            self.deiconify()
            self.init_ui()

        ctk.CTkButton(win, text="👉 Continuar", command=continuar).pack(pady=30)
        win.bind('<Return>', lambda event: continuar())

    def cerrar_sesion(self):
        """Cierra la sesión del usuario y vuelve a la ventana de login."""
        self.withdraw()
        for widget in self.winfo_children():
            widget.destroy()
        self.usuario = None
        self.is_admin = False
        self.login_usuario()

    def init_ui(self):
        """Inicializa la interfaz principal de la aplicación."""
        self.title(f"N.I.C.O.L.E – Sesión de {self.usuario}")
        self.state('zoomed')
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        try:
            self.brain_icon = ctk.CTkImage(Image.open("brain_icon.png"), size=(40, 40))
            self.help_icon = ctk.CTkImage(Image.open("help_icon.png"), size=(25, 25))
        except FileNotFoundError:
            self.brain_icon = None
            self.help_icon = None
            messagebox.showwarning("Iconos no encontrados", "Asegúrate de tener los iconos necesarios.")
        
        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        top_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top_frame, text="N.I.C.O.L.E", image=self.brain_icon, compound="left", font=ctk.CTkFont(size=22, weight="bold")).grid(row=0, column=0, padx=10, pady=10)

        self.busc = ctk.CTkEntry(top_frame, placeholder_text="🔍 Buscar en la tabla actual…", height=35)
        self.busc.grid(row=0, column=1, sticky="ew", padx=10)
        self.busc.bind("<KeyRelease>", self.buscar)

        frame_selector = ctk.CTkFrame(top_frame, fg_color="transparent")
        frame_selector.grid(row=0, column=2, padx=10)
        ctk.CTkLabel(frame_selector, text="Tabla:").pack(side="left", padx=(0, 5))
        tablas = obtener_tablas()
        self.combo = ctk.CTkOptionMenu(frame_selector, values=tablas if tablas else ["No hay tablas"], command=self.cargar_tabla, height=35)
        self.combo.pack(side="left")

        # Botón de Ayuda en la esquina superior derecha
        btn_help = ctk.CTkButton(top_frame, text="", image=self.help_icon, width=40, command=self.mostrar_ayuda)
        btn_help.grid(row=0, column=3, padx=10, pady=10, sticky="e")

        ttk_frame = ctk.CTkFrame(self)
        ttk_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        ttk_frame.grid_rowconfigure(0, weight=1)
        ttk_frame.grid_columnconfigure(0, weight=1)
        
        # Estilos de la tabla (ttk.Treeview)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2a2d2e", foreground="white", rowheight=25, fieldbackground="#343638", borderwidth=0)
        style.map('Treeview', background=[('selected', '#245d81')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat", font=('Calibri', 10, 'bold'))
        style.map("Treeview.Heading", background=[('active', '#343638')])

        self.tree = ttk.Treeview(ttk_frame, style="Treeview")
        vsb = ttk.Scrollbar(ttk_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(ttk_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<Double-1>", self.editar_celda)

        bottom_frame = ctk.CTkFrame(self)
        bottom_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=10)
        bottom_frame.grid_columnconfigure((1, 3), weight=1)
        
        # Botones de acción (Añadir, Borrar)
        frame_acciones = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_acciones.grid(row=0, column=0)
        self.btn_add = ctk.CTkButton(frame_acciones, text="➕ Añadir", command=self.abrir_form)
        self.btn_add.pack(side="left", padx=5)
        self.btn_del = ctk.CTkButton(frame_acciones, text="🗑️ Borrar", fg_color="#D32F2F", hover_color="#B71C1C", command=self.eliminar_registro)
        self.btn_del.pack(side="left", padx=5)

        # Botones de gestión (Recargar, Historial, Exportar)
        frame_gestion = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_gestion.grid(row=0, column=2)
        ctk.CTkButton(frame_gestion, text="🔄 Recargar", command=self.cargar_tabla_actual).pack(side="left", padx=5)
        ctk.CTkButton(frame_gestion, text="🕵️ Historial", command=self.ver_historial).pack(side="left", padx=5)
        ctk.CTkButton(frame_gestion, text="📦 Exportar", command=self.iniciar_exportacion).pack(side="left", padx=5)
        
        # Botones de sesión (Usuarios, Cambiar datos, Salir)
        frame_sesion = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        frame_sesion.grid(row=0, column=4)
        self.btn_user = ctk.CTkButton(frame_sesion, text="👤 Nuevos usuarios", command=self.registrar_usuario)
        self.btn_user.pack(side="left", padx=5)
        self.btn_change_data = ctk.CTkButton(frame_sesion, text="⚙️ Cambiar mis datos", command=self.cambiar_datos)
        self.btn_change_data.pack(side="left", padx=5)
        ctk.CTkButton(frame_sesion, text="🚪 Salir", command=self.cerrar_sesion).pack(side="left", padx=5)

        # Deshabilita los botones de administrador si el usuario no es admin
        if not self.is_admin:
            self.btn_del.configure(state="disabled")
            self.btn_user.configure(state="disabled")
            self.btn_add.configure(state="disabled")
            self.tree.unbind("<Double-1>")

        if tablas:
            self.combo.set(tablas[0])
            self.cargar_tabla(tablas[0])

    def iniciar_exportacion(self):
        """Inicia el proceso de exportación de datos a diferentes formatos."""
        if not hasattr(self, "table") or self.table == "No hay tablas":
            return mostrar_toast("Primero selecciona una tabla con datos.", self)
        
        file_types = [
            ("Archivos Excel", "*.xlsx"),
            ("Archivos PDF", "*.pdf"),
            ("Archivos CSV", "*.csv")
        ]
        
        filepath = filedialog.asksaveasfilename(
            filetypes=file_types,
            defaultextension=".xlsx",
            initialfile=f"export_{self.table}_{datetime.now():%Y-%m-%d}",
            title="Exportar datos de la tabla"
        )

        if not filepath:
            return

        if filepath.endswith('.xlsx'):
            self._exportar_excel(filepath)
        elif filepath.endswith('.csv'):
            self._exportar_csv(filepath)
        elif filepath.endswith('.pdf'):
            self._exportar_pdf(filepath)

    def _exportar_csv(self, filepath):
        """Función auxiliar para exportar datos a un archivo CSV."""
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(self.col)
                for item_id in self.tree.get_children():
                    writer.writerow(self.tree.item(item_id)['values'])
            messagebox.showinfo("Éxito", "CSV exportado con éxito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a CSV:\n{e}")

    def _exportar_excel(self, filepath):
        """Función auxiliar para exportar datos a un archivo Excel (.xlsx)."""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.table
            sheet.append(self.col)
            for item_id in self.tree.get_children():
                sheet.append(self.tree.item(item_id)['values'])
            workbook.save(filepath)
            messagebox.showinfo("Éxito", "Excel exportado con éxito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a Excel:\n{e}")

    def _exportar_pdf(self, filepath):
        """Función auxiliar para exportar datos a un archivo PDF."""
        try:
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, f'Reporte de Tabla: {self.table}', 0, 1, 'C')
            pdf.ln(5)

            pdf.set_font("Arial", 'B', 8)
            num_cols = len(self.col)
            page_width = pdf.w - 2 * pdf.l_margin
            col_width = page_width / num_cols if num_cols > 0 else 10

            pdf.set_fill_color(224, 224, 224)
            for header in self.col:
                pdf.cell(col_width, 8, str(header), 1, 0, 'C', fill=True)
            pdf.ln()

            pdf.set_font("Arial", '', 7)
            pdf.set_fill_color(255, 255, 255)
            fill = False
            for item_id in self.tree.get_children():
                row = self.tree.item(item_id)['values']
                for item in row:
                    pdf.cell(col_width, 8, str(item), 1, 0, 'L', fill=fill)
                pdf.ln()
                fill = not fill

            pdf.output(filepath)
            messagebox.showinfo("Éxito", "PDF exportado con éxito.", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a PDF:\n{e}")
            
    def registrar_usuario(self):
        """Muestra una ventana para que un administrador pueda crear nuevos usuarios."""
        if not self.is_admin: return

        win = ctk.CTkToplevel(self)
        win.title("➕ Nuevo Usuario")
        win.geometry("450x450")
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        ctk.CTkLabel(win, text="Nombre de Usuario:", wraplength=400).pack(pady=(20, 5))
        e_user = ctk.CTkEntry(win, width=350)
        e_user.pack(padx=20, fill="x")
        ctk.CTkLabel(win, text="Correo Electrónico:", wraplength=400).pack(pady=5)
        e_email = ctk.CTkEntry(win, width=350)
        e_email.pack(padx=20, fill="x")
        ctk.CTkLabel(win, text="Contraseña:", wraplength=400).pack(pady=5)
        e_pass_frame = PasswordEntry(win)
        e_pass_frame.pack(padx=20, fill="x")
        val_admin = ctk.CTkCheckBox(win, text="Otorgar privilegios de Administrador")
        val_admin.pack(pady=10)

        def guardar_user():
            """Guarda el nuevo usuario en la base de datos."""
            u = e_user.get().strip()
            p = e_pass_frame.get().strip()
            e = e_email.get().strip()
            
            if not u or not p or not e:
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)
            
            if not validar_email(e):
                return messagebox.showwarning("Correo inválido", "Formato de correo inválido.", parent=win)
            
            validation_error = validar_contrasena(p)
            if validation_error:
                return messagebox.showwarning("Contraseña débil", validation_error, parent=win)

            h = hashlib.sha256(p.encode()).hexdigest()
            esadm = val_admin.get()
            
            db = conectar_db()
            if not db: return
            
            try:
                cur = db.cursor()
                cur.execute("INSERT INTO usuarios_app (username, password_hash, email, es_admin) VALUES (%s, %s, %s, %s)", (u, h, e, esadm))
                db.commit()
                
                vals_dict = {"username": u, "email": e, "es_admin": esadm, "password_hash": "***"}
                registrar_historial(self.usuario, "INSERT", "usuarios_app", u, "", str(vals_dict))
                
                messagebox.showinfo("Éxito", "Usuario creado correctamente.", parent=win)
                win.destroy()
            except IntegrityError:
                messagebox.showerror("Error", "El usuario o correo ya existe.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error", f"No se pudo crear el usuario:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        ctk.CTkButton(win, text="💾 Guardar Usuario", command=guardar_user).pack(pady=20)

    def recuperar_contrasena(self):
        """Muestra la ventana para recuperar la contraseña del usuario."""
        win = ctk.CTkToplevel(self)
        win.title("🔑 Recuperar Contraseña")
        win.geometry("350x300")
        win.resizable(False, False)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)
        
        frame1 = ctk.CTkFrame(win)
        frame1.pack(fill="both", expand=True, padx=20, pady=20)
        ctk.CTkLabel(frame1, text="Ingresa tu correo o nombre de usuario:").pack(pady=(20, 5))
        entry = ctk.CTkEntry(frame1)
        entry.pack(fill="x")

        def enviar_codigo():
            """Envía un código de recuperación por correo electrónico."""
            identificador = entry.get().strip()
            if not identificador: return messagebox.showwarning("Campo requerido", "Por favor, ingresa tu usuario o correo.", parent=win)

            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                cur.execute("SELECT email, username FROM usuarios_app WHERE username=%s OR email=%s", (identificador, identificador))
                res = cur.fetchone()
                if not res:
                    return messagebox.showwarning("No encontrado", "Usuario o correo no encontrado.", parent=win)
                
                email = res[0]
                username = res[1]
                codigo = "".join(random.choices("0123456789", k=6))
                
                if enviar_email_cod(email, codigo):
                    expiracion = datetime.now() + timedelta(minutes=15)
                    cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (username,))
                    cur.execute("INSERT INTO recuperacion_codigos (username, email, codigo, expiracion) VALUES (%s, %s, %s, %s)", (username, email, codigo, expiracion))
                    db.commit()
                    messagebox.showinfo("Código enviado", "Código enviado. Revisa tu correo.", parent=win)
                    frame1.pack_forget()
                    frame2.pack(fill="both", expand=True, padx=20, pady=20)
                
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurrió un error:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        ctk.CTkButton(frame1, text="Enviar Código", command=enviar_codigo).pack(pady=20)

        frame2 = ctk.CTkFrame(win)
        ctk.CTkLabel(frame2, text="Ingresa el código y la nueva contraseña:", wraplength=300).pack(pady=(20, 5))
        entry_codigo = ctk.CTkEntry(frame2, placeholder_text="Código de 6 dígitos")
        entry_codigo.pack(fill="x", padx=20)
        ctk.CTkLabel(frame2, text="Nueva Contraseña:").pack(pady=5)
        entry_new_pass_frame = PasswordEntry(frame2)
        entry_new_pass_frame.pack(fill="x", padx=20)

        def actualizar_contrasena():
            """Actualiza la contraseña del usuario si el código es válido."""
            codigo = entry_codigo.get().strip()
            new_pass = entry_new_pass_frame.get().strip()
            identificador = entry.get().strip()

            if not codigo or not new_pass:
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)

            validation_error = validar_contrasena(new_pass)
            if validation_error:
                return messagebox.showwarning("Contraseña débil", validation_error, parent=win)

            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                cur.execute("SELECT username FROM recuperacion_codigos WHERE (username=%s OR email=%s) AND codigo=%s AND expiracion > NOW()", (identificador, identificador, codigo))
                res = cur.fetchone()

                if res:
                    h = hashlib.sha256(new_pass.encode()).hexdigest()
                    cur.execute("UPDATE usuarios_app SET password_hash=%s WHERE username=%s", (h, res[0]))
                    cur.execute("DELETE FROM recuperacion_codigos WHERE username=%s", (res[0],))
                    db.commit()
                    
                    registrar_historial("Sistema", "UPDATE", "usuarios_app", res[0], "password_hash: ****", "password_hash: ****")

                    messagebox.showinfo("Éxito", "Contraseña actualizada con éxito.", parent=win)
                    win.destroy()
                else:
                    messagebox.showwarning("Código inválido", "Código inválido o expirado.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error de BD", f"Ocurrió un error:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        ctk.CTkButton(frame2, text="Actualizar Contraseña", command=actualizar_contrasena).pack(pady=20)

    def cambiar_datos(self):
        """Muestra una ventana para que el usuario cambie su correo o contraseña."""
        win = ctk.CTkToplevel(self)
        win.title("⚙️ Cambiar Mis Datos")
        win.geometry("450x400")
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        db = conectar_db()
        if not db: return
        email_actual = ""
        try:
            cur = db.cursor()
            cur.execute("SELECT email FROM usuarios_app WHERE username=%s", (self.usuario,))
            result = cur.fetchone()
            if result:
                email_actual = result[0]
        except MySQLError:
            email_actual = ""
        finally:
            if db.open: cur.close(); db.close()

        ctk.CTkLabel(win, text=f"Usuario: {self.usuario}", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(20, 10))
        ctk.CTkLabel(win, text="Nuevo Correo Electrónico: (dejar igual si no se quiere cambiar el correo electrónico)", wraplength=400).pack(pady=(10, 5))
        e_email = ctk.CTkEntry(win, width=350)
        e_email.insert(0, email_actual)
        e_email.pack(padx=20, fill="x")
        ctk.CTkLabel(win, text="Nueva Contraseña: (Deja en blanco para no cambiarla)", wraplength=400).pack(pady=5)
        e_pass_frame = PasswordEntry(win)
        e_pass_frame.pack(padx=20, fill="x")

        def actualizar_datos():
            """Actualiza el correo y/o la contraseña del usuario en la base de datos."""
            new_email = e_email.get().strip()
            new_pass = e_pass_frame.get().strip()

            if not validar_email(new_email):
                return messagebox.showwarning("Correo inválido", "Formato de correo inválido.", parent=win)

            db = conectar_db()
            if not db: return
            try:
                cur = db.cursor()
                if new_pass:
                    validation_error = validar_contrasena(new_pass)
                    if validation_error:
                        return messagebox.showwarning("Contraseña débil", validation_error, parent=win)
                    h = hashlib.sha256(new_pass.encode()).hexdigest()
                    cur.execute("UPDATE usuarios_app SET email=%s, password_hash=%s WHERE username=%s", (new_email, h, self.usuario))
                    registrar_historial(self.usuario, "UPDATE", "usuarios_app", self.usuario, f"email: {email_actual}, password:***", f"email: {new_email}, password:***")
                else:
                    cur.execute("UPDATE usuarios_app SET email=%s WHERE username=%s", (new_email, self.usuario))
                    registrar_historial(self.usuario, "UPDATE", "usuarios_app", self.usuario, f"email: {email_actual}", f"email: {new_email}")
                
                db.commit()
                messagebox.showinfo("Éxito", "Datos actualizados correctamente.", parent=win)
                win.destroy()
            except IntegrityError:
                messagebox.showerror("Error", "El correo ya está en uso.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error", f"No se pudieron actualizar los datos:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        ctk.CTkButton(win, text="💾 Guardar Cambios", command=actualizar_datos).pack(pady=20)

    def cargar_tabla(self, tabla):
        """Carga los datos de la tabla seleccionada en el Treeview."""
        if tabla == "No hay tablas":
            self.tree.delete(*self.tree.get_children())
            self.tree["columns"] = []
            return
        
        # Muestra un indicador de carga
        self.loading_label = ctk.CTkLabel(self, text="Cargando...", font=ctk.CTkFont(size=16))
        self.loading_label.grid(row=2, column=0, sticky="nsew")
        self.update_idletasks()
            
        self.tree.delete(*self.tree.get_children())
        self.col, rows = obtener_datos(tabla)
        
        # Elimina el indicador de carga
        self.loading_label.grid_forget()
        
        if not self.col:
            return

        self.table = tabla
        self.tree["columns"] = self.col
        self.tree["show"] = "headings"
        for c in self.col:
            self.tree.heading(c, text=c)
            self.tree.column(c, anchor="w", width=150, minwidth=100)
        for r in rows:
            self.tree.insert("", "end", values=r)

    def cargar_tabla_actual(self):
        """Recarga la tabla que se está mostrando actualmente."""
        if hasattr(self, "table"):
            self.cargar_tabla(self.table)

    def buscar(self, ev=None):
        """Filtra los datos de la tabla según el texto de búsqueda."""
        if not hasattr(self, "table"): return
        txt = self.busc.get().lower()
        self.tree.delete(*self.tree.get_children())
        _, rows = obtener_datos(self.table)
        for r in rows:
            if any(txt in str(v).lower() for v in r):
                self.tree.insert("", "end", values=r)

    def abrir_form(self):
        """Abre un formulario para añadir un nuevo registro a la tabla actual."""
        if not hasattr(self, "table"):
            return mostrar_toast("Selecciona una tabla primero", self)

        win = ctk.CTkToplevel(self)
        win.title(f"➕ Agregar registro a {self.table}")
        win.geometry("500x500")
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)

        scroll = ctk.CTkScrollableFrame(win)
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        cols, _ = obtener_datos(self.table)
        cols_no_id = [c for c in cols if c.lower() != "id"]
        entries = {}

        for c in cols_no_id:
            ctk.CTkLabel(scroll, text=f"{c}:", wraplength=400).pack(anchor="w", padx=5, pady=(10, 0))
            if "fecha" in c.lower():
                ent = ctk.CTkEntry(scroll, width=400)
                ent.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                ent.configure(state="readonly")
                ent.pack(fill="x", padx=5)
                entries[c] = ent
            elif "pass" in c.lower() and "password_hash" not in c.lower():
                ent = PasswordEntry(scroll, width=400)
                ent.pack(fill="x", padx=5)
                entries[c] = ent
            else:
                ent = ctk.CTkEntry(scroll, width=400)
                ent.pack(fill="x", padx=5)
                entries[c] = ent
        
        def insertar():
            """Inserta el nuevo registro en la base de datos."""
            vals_dict = {}
            for c in cols_no_id:
                v = entries[c].get()
                if "pass" in c.lower() and "password_hash" not in c.lower():
                    validation_error = validar_contrasena(v)
                    if validation_error:
                        return messagebox.showwarning("Contraseña débil", validation_error, parent=win)
                    v = hashlib.sha256(v.encode()).hexdigest()
                vals_dict[c] = v

            vals = list(vals_dict.values())
            
            if any(not str(x).strip() for x in vals):
                return messagebox.showwarning("Campos incompletos", "Completa todos los campos.", parent=win)
            
            db = conectar_db()
            if not db: return

            try:
                cur = db.cursor()
                # Se construye la consulta de forma segura
                query = f"INSERT INTO `{self.table}` ({','.join([f'`{c}`' for c in cols_no_id])}) VALUES ({','.join(['%s'] * len(vals))})"
                cur.execute(query, vals)
                rid = cur.lastrowid
                db.commit()
                
                if "password_hash" in vals_dict:
                    vals_dict["password_hash"] = "***"
                
                registrar_historial(self.usuario, "INSERT", self.table, rid, "", str(vals_dict))
                win.destroy()
                self.cargar_tabla_actual()
                messagebox.showinfo("Éxito", "Registro insertado.", parent=win)
            except MySQLError as err:
                messagebox.showerror("Error al Insertar", f"No se pudo guardar:\n{err}", parent=win)
            finally:
                if db.open: cur.close(); db.close()

        ctk.CTkButton(scroll, text="💾 Insertar Registro", command=insertar).pack(pady=20, padx=5)

    def eliminar_registro(self):
        """Elimina un registro seleccionado del Treeview y la base de datos."""
        sel = self.tree.focus()
        if not sel: return

        r = self.tree.item(sel)["values"]
        pk_col_name = self.col[0]
        pk_value = r[0]
        before = str(dict(zip(self.col, r)))

        # Se pide confirmación al usuario antes de borrar (mejora de UX)
        if messagebox.askyesno("Confirmar Borrado", f"¿Seguro que quieres eliminar el registro con {pk_col_name} = {pk_value}?"):
            db = conectar_db()
            if not db: return
            
            try:
                cur = db.cursor()
                # Se construye la consulta de forma segura
                cur.execute(f"DELETE FROM `{self.table}` WHERE `{pk_col_name}`=%s", (pk_value,))
                db.commit()
                
                registrar_historial(self.usuario, "DELETE", self.table, pk_value, before, "")
                self.cargar_tabla_actual()
                messagebox.showinfo("Éxito", "Registro eliminado.", parent=self)
            except IntegrityError as e:
                # Manejo de error de clave foránea
                if e.args[0] == 1451:
                    match = re.search(r"fails \(`\w+`\.`(.+?)`\)", e.args[1])
                    tabla_dependiente = match.group(1) if match else "otra tabla"
                    msg = (f"No se puede eliminar. Este registro está siendo usado por la tabla '{tabla_dependiente}'.\n\n"
                           f"Debes eliminar o actualizar los registros dependientes primero.")
                    messagebox.showerror("Error de Integridad", msg)
                else:
                    messagebox.showerror("Error de BD", f"No se pudo eliminar:\n{e}")
            except MySQLError as err:
                messagebox.showerror("Error de Borrado", f"Ocurrió un error:\n{err}")
            finally:
                if db.open: cur.close(); db.close()

    def editar_celda(self, ev):
        """Permite editar el valor de una celda directamente en el Treeview."""
        if not self.is_admin: return
            
        sel = self.tree.identify_row(ev.y)
        cid = self.tree.identify_column(ev.x)
        if not sel or not cid: return

        col_index = int(cid.replace("#", "")) - 1
        col_name = self.col[col_index]
        pk_col_name = self.col[0]
        pk_value = self.tree.item(sel)["values"][0]

        if col_name.lower() == pk_col_name.lower():
            return mostrar_toast(f"La clave primaria ({pk_col_name}) no se puede editar.", self)
            
        old_val = self.tree.item(sel)["values"][col_index]
        
        x, y, w, h = self.tree.bbox(sel, cid)
        
        is_password_field = "pass" in col_name.lower() and "hash" not in col_name.lower()
        
        if is_password_field:
            ent = PasswordEntry(self.tree, width=w, height=h)
            ent.place(x=x, y=y)
            ent.entry.insert(0, old_val)
        else:
            ent = ctk.CTkEntry(self.tree, width=w, height=h)
            ent.place(x=x, y=y)
            ent.insert(0, old_val)
        
        ent.focus()

        def save(event):
            """Guarda el nuevo valor de la celda en la base de datos."""
            try:
                new_val = ent.get()
                ent.destroy()
            except TclError:
                return 
            
            if new_val == old_val: return

            final_val = new_val
            if is_password_field:
                validation_error = validar_contrasena(new_val)
                if validation_error:
                    return messagebox.showwarning("Contraseña débil", validation_error, parent=self)
                final_val = hashlib.sha256(new_val.encode()).hexdigest()

            db = conectar_db()
            if not db: return
            
            try:
                cur = db.cursor()
                # Se construye la consulta de forma segura
                cur.execute(f"UPDATE `{self.table}` SET `{col_name}`=%s WHERE `{pk_col_name}`=%s", (final_val, pk_value))
                db.commit()
                
                # Se registran los cambios para el historial
                antes = f"{col_name}: {old_val}"
                despues = f"{col_name}: {final_val}"
                if "pass" in col_name.lower():
                    antes = f"{col_name}: ***"
                    despues = f"{col_name}: ***"

                registrar_historial(self.usuario, "UPDATE", self.table, pk_value, antes, despues)
                self.cargar_tabla_actual()
                
                # Feedback visual con un "toast" (mejora de UX)
                mostrar_toast("Campo actualizado.", self)
            except MySQLError as err:
                messagebox.showerror("Error al Actualizar", f"No se pudo guardar:\n{err}")
                self.cargar_tabla_actual()
            finally:
                if db.open: cur.close(); db.close()

        ent.bind("<Return>", save)
        ent.bind("<FocusOut>", lambda e: ent.destroy())

    def ver_historial(self):
        """Muestra una ventana con el historial completo de cambios en la base de datos."""
        win = ctk.CTkToplevel(self)
        win.title("📋 Historial de Cambios")
        win.geometry('1000x500')
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)
        
        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        style = ttk.Style(frame)
        style.theme_use("default")
        style.configure("Treeview", background="#2a2d2e", foreground="white", rowheight=25, fieldbackground="#343638", borderwidth=0)
        style.map('Treeview', background=[('selected', '#245d81')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat", font=('Calibri', 10, 'bold'))
        style.map("Treeview.Heading", background=[('active', '#343638')])
        
        tv = ttk.Treeview(frame, style="Treeview")
        tv["columns"] = ("ID", "Usuario", "Fecha", "Acción", "Tabla", "ID Reg.", "Valores Anteriores", "Valores Nuevos")
        tv.column("#0", width=0, stretch="no")

        tv.column("ID", width=50, minwidth=50, stretch="no")
        tv.column("Usuario", width=120, minwidth=100, stretch="no")
        tv.column("Fecha", width=150, minwidth=130, stretch="no")
        tv.column("Acción", width=100, minwidth=80, stretch="no")
        tv.column("Tabla", width=120, minwidth=100, stretch="no")
        tv.column("ID Reg.", width=80, minwidth=70, stretch="no")
        tv.column("Valores Anteriores", anchor="w", stretch="yes", minwidth=400, width=500)
        tv.column("Valores Nuevos", anchor="w", stretch="yes", minwidth=400, width=500)

        for col in tv["columns"]:
            tv.heading(col, text=col, anchor="w")
        
        db = conectar_db()
        if not db: return

        try:
            cur = db.cursor()
            cur.execute("SELECT id, usuario, fecha, accion, tabla, registro_id, valores_antes, valores_despues FROM historial ORDER BY fecha DESC")
            for r in cur.fetchall():
                tv.insert("", "end", values=r)
        except MySQLError as err:
            messagebox.showerror("Error", f"No se pudo cargar el historial:\n{err}", parent=win)
        finally:
            if db.open: cur.close(); db.close()
        
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tv.xview)
        tv.configure(xscrollcommand=hsb.set)
        
        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

    def mostrar_ayuda(self):
        """Muestra una ventana de ayuda y preguntas frecuentes."""
        win = ctk.CTkToplevel(self)
        win.title("❓ Ayuda y Preguntas Frecuentes")
        win.geometry("600x400")
        win.resizable(True, True)
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.after(10, win.grab_set)
        
        scroll_frame = ctk.CTkScrollableFrame(win)
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # El contenido de la ayuda varía si el usuario es administrador
        if self.is_admin:
            ctk.CTkLabel(scroll_frame, text="Ayuda para Administradores", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(10, 5))
            faq = {
                "¿Cómo añado un nuevo registro?": "Selecciona una tabla y haz clic en '➕ Añadir'. Completa el formulario y haz clic en 'Guardar'.",
                "¿Cómo edito un dato?": "Haz doble clic sobre la celda que deseas modificar. Escribe el nuevo valor y presiona 'Enter'.",
                "¿Por qué no puedo eliminar un registro?": "Es posible que el registro esté siendo usado como clave foránea en otra tabla. Debes eliminar primero los registros dependientes.",
                "¿Cómo creo un nuevo usuario o cambio sus datos?": "Haz clic en '👤 Nuevos usuarios' para crear una nueva cuenta. Puedes usar '⚙️ Cambiar mis datos' para modificar tu información.",
                "¿Qué información me muestra el historial?": "El 'Historial de Cambios' registra todas las acciones de inserción, actualización y eliminación de datos, incluyendo quién las realizó y cuándo. Los datos sensibles como contraseñas y claves de usuarios están enmascarados para mayor seguridad.",
                "Mi contraseña no es aceptada, ¿por qué?": "Las contraseñas deben tener un mínimo de 8 caracteres, al menos una letra mayúscula y un carácter especial.",
                "Olvidé mi contraseña, ¿qué hago?": "En la ventana de login, haz clic en 'Olvidé mi contraseña' y sigue las instrucciones para recibir un código de recuperación por correo.",
                "No puedo conectar a la base de datos": "Verifica que el servidor MySQL esté corriendo y que la configuración en tu archivo '.env' sea correcta (host, usuario, contraseña, etc.)."
            }
        else:
            ctk.CTkLabel(scroll_frame, text="Ayuda para Usuarios", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(10, 5))
            faq = {
                "¿Cómo añado un nuevo registro?": "Selecciona una tabla y haz clic en '➕ Añadir'. Completa el formulario y haz clic en 'Guardar'.",
                "¿Cómo edito un dato?": "Haz doble clic sobre la celda que deseas modificar. Escribe el nuevo valor y presiona 'Enter'.",
                "¿Por qué no puedo eliminar un registro?": "Tu nivel de usuario no tiene permisos para borrar registros. Contacta a un administrador para realizar esta acción.",
                "Mi contraseña no es aceptada, ¿por qué?": "Las contraseñas deben tener un mínimo de 8 caracteres, al menos una letra mayúscula y un carácter especial.",
                "Olvidé mi contraseña, ¿qué hago?": "En la ventana de login, haz clic en 'Olvidé mi contraseña' y sigue las instrucciones para recibir un código de recuperación por correo."
            }
        
        for pregunta, respuesta in faq.items():
            ctk.CTkLabel(scroll_frame, text=f"**{pregunta}**", font=ctk.CTkFont(size=14, weight="bold"), wraplength=550, justify="left").pack(anchor="w", pady=(10, 0))
            ctk.CTkLabel(scroll_frame, text=respuesta, font=ctk.CTkFont(size=12), wraplength=550, justify="left").pack(anchor="w", pady=(0, 5))

# -----------------------------------------------
# PUNTO DE ENTRADA DE LA APLICACIÓN
# -----------------------------------------------
if __name__ == "__main__":
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    app = App()
    app.mainloop()